
import csv
import io
import json
import logging
import os
import re
import sqlite3
import unicodedata
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from threading import Lock
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlencode, urljoin
from urllib.request import Request, urlopen

from fastapi import Body, FastAPI, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, StreamingResponse
from openpyxl import load_workbook

APP_TITLE = "Gestion Affaire - Finance"
DEFAULT_WORKBOOK_PATH = r"\\192.168.10.100\03 - finances\10 - TABLEAU DE BORD\01 - TABLEAU ACTIVITEE\TABLEAU ACTIVITEE.xlsx"
DEFAULT_SHEET_NAME = "AFFAIRES 2026"
DEFAULT_CACHE_FILE = "finance_cache.json"

WORKBOOK_PATH = os.getenv("ACTIVITE_XLSX_PATH", DEFAULT_WORKBOOK_PATH)
SHEET_NAME = os.getenv("ACTIVITE_SHEET_NAME", DEFAULT_SHEET_NAME)
CACHE_FILE = os.getenv("FINANCE_CACHE_FILE", DEFAULT_CACHE_FILE)
EXPECTED_SCHEMA_VERSION = "finance_affaires_dataset_v4"
TEMPO_LOGO_PATH = os.getenv("TEMPO_LOGO_PATH", r"\\192.168.10.100\02 - affaires\02.2 - SYNTHESE\ZZ - METRONOME\Content\T logo.png")
METRONOME_BASE_PATH = os.getenv("METRONOME_BASE_PATH", r"\\192.168.10.100\02 - affaires\02.2 - SYNTHESE\ZZ - METRONOME")
POINTAGE_STORE_FILE = Path(os.getenv("POINTAGE_STORE_FILE", "pointage_store.json"))
BOOND_CACHE_DB_PATH = Path(os.getenv("BOOND_CACHE_DB_PATH", "boond_cache.sqlite3"))
METRONOME_FILES = {
    "projects": "Projects.csv",
    "entries": "Entries (Tasks & Memos).csv",
    "meetings": "Meetings.csv",
    "areas": "Areas.csv",
    "packages": "Packages.csv",
    "companies": "Companies.csv",
    "users": "Users.csv",
    "comments": "Comments.csv",
    "documents": "Documents.csv",
}

METRONOME_COLUMN_ALIASES = {
    "project_title_entries": ["Project/Title", "Project/Title (dev)", "Project"],
    "project_title_projects": ["Title", "Name"],
    "project_name_projects": ["Name", "Title"],
    "project_start": ["Start Date", "StartDate", "Start"],
    "project_end": ["End Date", "EndDate", "End"],
    "project_image": ["Image", "Image URL", "Project Image"],
    "project_description": ["Description", "Start Sentence"],
    "entry_done_date": ["Done Date", "DoneDate", "Completed Date", "CompletedDate", "Completed/Declared End", "ClosedDate", "UpdatedAt"],
    "entry_category": ["Category/Name to display", "Category"],
    "entry_project_id": ["Project/ID", "Project"],
    "entry_meeting_id": ["Meeting/ID", "Meeting"],
    "entry_owner_id": ["Owner for Tasks/ID", "Owner/ID", "Assignee"],
    "entry_created_by_id": ["Created by/ID", "CreatedBy/ID"],
    "entry_comment_editor_id": ["Comment for Tasks/Editor ID"],
    "entry_completed_reporting_user_id": ["Completed/Reporting User ID"],
    "entry_completed_declared_user_id": ["Completed/Declared User ID"],
    "entry_task_package_id": ["Packages/ID for Task", "Package", "Package/ID"],
    "entry_memo_package_ids": ["Packages/IDs for Memos"],
    "entry_area_ids": ["Areas/IDs", "Area/ID", "Area"],
    "entry_deadline": ["Deadline & Status for Tasks/Deadline", "DueDate"],
    "entry_status_id": ["Deadline & Status for Tasks/Status ID", "Status ID"],
    "entry_status_label": ["Deadline & Status for Tasks/Status Emoji + Text", "Status"],
    "entry_comment_text": ["Comment for Tasks/Text", "Comment", "Text"],
    "comment_memo_id": ["Memo/ID", "Entry", "Memo"],
    "comment_owner_id": ["Owner/ID", "Owner"],
    "comment_date": ["Date", "Created", "Creation Date", "UpdatedAt"],
    "meeting_date": ["Date", "Meeting Date"],
    "meeting_attending_company_ids": ["Companies/Attending IDs"],
    "meeting_missing_company_ids": ["Companies/Missing IDs", "Companies/Missing Calculated IDs (dev)"],
    "package_company_id": ["Company/ID", "Company"],
    "package_project_id": ["Project/ID"],
    "area_project_id": ["Project/ID"],
    "company_collaborators_ids": ["Collaborators/IDs"],
    "documents_project_id": ["Project/ID"],
    "documents_meeting_id": ["Meeting/ID"],
}

MONTHS = [
    "janvier", "fevrier", "mars", "avril", "mai", "juin",
    "juillet", "aout", "septembre", "octobre", "novembre", "decembre",
]
MONTH_LABELS = {
    "janvier": "Janv.", "fevrier": "Févr.", "mars": "Mars", "avril": "Avr.",
    "mai": "Mai", "juin": "Juin", "juillet": "Juil.", "aout": "Août",
    "septembre": "Sept.", "octobre": "Oct.", "novembre": "Nov.", "decembre": "Déc.",
}
COLUMN_ORDER = [
    "client", "affaire", "tag", "numero", "delai_reglement_jours", "commande_ht",
    "facturation_cumulee_2017", "facturation_cumulee_2018", "facturation_cumulee_2021",
    "facturation_cumulee_2022", "facturation_cumulee_2023", "facturation_cumulee_2024",
    "facturation_cumulee_2025", "facturation_cumulee_2026", "reste_a_facturer",
    "janvier_previsionnel", "janvier_facture", "fevrier_previsionnel", "fevrier_facture",
    "mars_previsionnel", "mars_facture", "avril_previsionnel", "avril_facture",
    "mai_previsionnel", "mai_facture", "juin_previsionnel", "juin_facture",
    "juillet_previsionnel", "juillet_facture", "aout_previsionnel", "aout_facture",
    "septembre_previsionnel", "septembre_facture", "octobre_previsionnel", "octobre_facture",
    "novembre_previsionnel", "novembre_facture", "decembre_previsionnel", "decembre_facture",
    "total_previsionnel", "total_facture",
]

logger = logging.getLogger("finance_cockpit")
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(name)s | %(message)s")

app = FastAPI(title=APP_TITLE)




def maybe_load_dotenv() -> None:
    """Charge .env si python-dotenv est disponible (sans dépendance obligatoire)."""
    dotenv_spec = __import__("importlib.util").util.find_spec("dotenv")
    if dotenv_spec is None:
        return
    dotenv_module = __import__("importlib").import_module("dotenv")
    load_dotenv = getattr(dotenv_module, "load_dotenv", None)
    if callable(load_dotenv):
        load_dotenv()


maybe_load_dotenv()


class BoondService:
    """Service BOOND : appels API, cache SQLite et agrégation imputations."""

    IMPORTANT_TOKENS = {"115", "CDG", "MDZ", "PASSY", "KENNEDY", "VALHUBERT", "PICPUS", "CONDORCET"}

    def __init__(self, db_path: Path) -> None:
        self.db_path = Path(db_path)
        self._cache_lock = Lock()
        self.manual_daily_rates_by_project: Dict[str, float] = {}

        self.base_url = clean_text(os.getenv("BOOND_BASE_URL"))
        self.client_token = clean_text(os.getenv("BOOND_CLIENT_TOKEN"))
        self.client_key = clean_text(os.getenv("BOOND_CLIENT_KEY"))
        self.user_token = clean_text(os.getenv("BOOND_USER_TOKEN"))
        self.ensure_boond_cache_table()

    def _validate_config(self) -> None:
        missing = [
            k for k, v in {
                "BOOND_BASE_URL": self.base_url,
                "BOOND_CLIENT_TOKEN": self.client_token,
                "BOOND_CLIENT_KEY": self.client_key,
                "BOOND_USER_TOKEN": self.user_token,
            }.items() if not v
        ]
        if missing:
            raise RuntimeError(f"[BOOND] Variables .env manquantes: {', '.join(missing)}")

    def boond_headers(self) -> Dict[str, str]:
        self._validate_config()
        auth = f"{self.client_token}:{self.client_key}:{self.user_token}"
        return {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "X-Jwt-Client-Boondmanager": auth,
        }

    def boond_get(self, path: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        self._validate_config()
        base = self.base_url.rstrip("/") + "/"
        path_clean = path.lstrip("/")
        url = urljoin(base, path_clean)
        if params:
            url = f"{url}?{urlencode(params, doseq=True)}"
        req = Request(url, headers=self.boond_headers(), method="GET")
        try:
            with urlopen(req, timeout=30) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except Exception as exc:
            logger.exception("[BOOND] Erreur GET %s", url)
            raise RuntimeError(f"Erreur BOOND sur {path}: {exc}")

    def _db(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def ensure_boond_cache_table(self) -> None:
        with self._cache_lock:
            with self._db() as conn:
                conn.execute(
                    """
                    CREATE TABLE IF NOT EXISTS boond_api_cache (
                        cache_key TEXT PRIMARY KEY,
                        payload TEXT NOT NULL,
                        expires_at INTEGER NOT NULL,
                        updated_at INTEGER NOT NULL
                    )
                    """
                )
                conn.commit()

    def get_api_cache(self, cache_key: str) -> Optional[Dict[str, Any]]:
        now_ts = int(datetime.now().timestamp())
        with self._db() as conn:
            row = conn.execute(
                "SELECT payload, expires_at FROM boond_api_cache WHERE cache_key = ?",
                (cache_key,),
            ).fetchone()
        if not row:
            return None
        if int(row["expires_at"]) < now_ts:
            return None
        return json.loads(row["payload"])

    def set_api_cache(self, cache_key: str, payload: Dict[str, Any], ttl_seconds: int) -> None:
        now_ts = int(datetime.now().timestamp())
        expires_at = now_ts + int(ttl_seconds)
        with self._db() as conn:
            conn.execute(
                """
                INSERT INTO boond_api_cache(cache_key, payload, expires_at, updated_at)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(cache_key) DO UPDATE SET
                    payload = excluded.payload,
                    expires_at = excluded.expires_at,
                    updated_at = excluded.updated_at
                """,
                (cache_key, json.dumps(payload, ensure_ascii=False), expires_at, now_ts),
            )
            conn.commit()

    def get_times_report_cached(self, times_report_id: str, ttl_seconds: int = 86400) -> Tuple[Dict[str, Any], str]:
        cache_key = f"boond:times-report:{times_report_id}"
        cached = self.get_api_cache(cache_key)
        if cached is not None:
            return cached, "cache"
        payload = self.boond_get(f"/times-reports/{times_report_id}")
        self.set_api_cache(cache_key, payload, ttl_seconds=ttl_seconds)
        return payload, "api"

    @staticmethod
    def extract_projects_from_times_report(report: Dict[str, Any]) -> List[Dict[str, str]]:
        # Chaîne fiable validée: workplaces-times -> times-report -> times -> delivery.project
        projects: Dict[str, Dict[str, str]] = {}
        rel_times = ((report.get("data") or {}).get("relationships") or {}).get("times") or {}
        items = rel_times.get("data") or []
        for item in items:
            attrs = item.get("attributes") or {}
            delivery = attrs.get("delivery") or {}
            project = delivery.get("project") or {}
            pid = clean_text(project.get("id"))
            pref = clean_text(project.get("reference"))
            if pid:
                projects[pid] = {"project_id": pid, "project_reference": pref}
        return list(projects.values())

    @staticmethod
    def normalize_text(value: Any) -> str:
        txt = clean_text(value).upper()
        txt = txt.replace("_", " ").replace("-", " ")
        txt = re.sub(r"[^A-Z0-9\s]", " ", txt)
        return re.sub(r"\s+", " ", txt).strip()

    def score_project_match(self, project_name: str, boond_ref: str) -> Tuple[int, str]:
        q = self.normalize_text(project_name)
        r = self.normalize_text(boond_ref)
        if not q or not r:
            return 0, "empty"
        if q == r:
            return 100, "exact_reference"
        if q in r or r in q:
            return 93, "token_contains"
        q_tokens = set(q.split())
        r_tokens = set(r.split())
        if not q_tokens or not r_tokens:
            return 0, "empty_tokens"
        overlap = q_tokens & r_tokens
        base = int((len(overlap) / max(1, len(q_tokens))) * 80)
        bonus = sum(8 for t in overlap if t in self.IMPORTANT_TOKENS)
        return min(99, base + bonus), "token_overlap"

    def find_best_boond_project_match(self, project_name: str, boond_projects: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
        best: Optional[Dict[str, Any]] = None
        for p in boond_projects:
            score, matched_on = self.score_project_match(project_name, p.get("project_reference"))
            if score < 55:
                continue
            candidate = {
                "project_id": clean_text(p.get("project_id")),
                "project_reference": clean_text(p.get("project_reference")),
                "match_score": score,
                "matched_on": matched_on,
            }
            if best is None or int(candidate["match_score"]) > int(best["match_score"]):
                best = candidate
        return best

    def _fetch_all_workplaces_times(self) -> List[Dict[str, Any]]:
        # Une seule collecte de workplaces-times, puis enrichissement avec times-reports en cache.
        payload = self.boond_get("/workplaces-times")
        data = payload.get("data") or []
        return data if isinstance(data, list) else []

    def build_boond_imputation_index(self, refresh: bool = False) -> Dict[str, Any]:
        cache_key = "boond:index:projects:imputations"
        if not refresh:
            cached = self.get_api_cache(cache_key)
            if cached is not None:
                return cached

        rows = self._fetch_all_workplaces_times()
        structured_rows = []
        unique_reports = set()
        for row in rows:
            attrs = row.get("attributes") or {}
            rel = row.get("relationships") or {}
            tr_id = clean_text((((rel.get("timesReport") or {}).get("data") or {}).get("id")))
            duration = clean_number(attrs.get("duration"))
            start_date = clean_text(attrs.get("startDate"))
            if tr_id:
                unique_reports.add(tr_id)
            structured_rows.append({"times_report_id": tr_id, "duration": duration, "start_date": start_date})

        report_to_projects: Dict[str, List[Dict[str, str]]] = {}
        loaded_api = 0
        loaded_cache = 0
        for tr_id in unique_reports:
            report_payload, source = self.get_times_report_cached(tr_id)
            if source == "api":
                loaded_api += 1
            else:
                loaded_cache += 1
            report_to_projects[tr_id] = self.extract_projects_from_times_report(report_payload)

        projects_index: Dict[str, Dict[str, Any]] = {}
        for wr in structured_rows:
            tr_projects = report_to_projects.get(wr["times_report_id"], [])
            for project in tr_projects:
                pid = project["project_id"]
                pref = project.get("project_reference") or ""
                if pid not in projects_index:
                    projects_index[pid] = {
                        "project_id": pid,
                        "project_reference": pref,
                        "total_days": 0.0,
                        "by_month": defaultdict(float),
                    }
                projects_index[pid]["total_days"] += wr["duration"]
                month = clean_text(wr["start_date"])[:7]
                if re.match(r"^\d{4}-\d{2}$", month):
                    projects_index[pid]["by_month"][month] += wr["duration"]

        serializable_projects = []
        for item in projects_index.values():
            serializable_projects.append({
                "project_id": item["project_id"],
                "project_reference": item["project_reference"],
                "total_days": round(item["total_days"], 2),
                "by_month": {k: round(v, 2) for k, v in sorted(item["by_month"].items())},
            })

        result = {
            "projects": serializable_projects,
            "meta": {
                "workplaces_rows": len(structured_rows),
                "unique_times_reports": len(unique_reports),
                "times_reports_loaded_from_api": loaded_api,
                "times_reports_loaded_from_cache": loaded_cache,
                "generated_at": now_iso(),
            },
        }
        self.set_api_cache(cache_key, result, ttl_seconds=1800)
        return result

    def get_project_imputation_summary(self, project_name: str, refresh: bool = False) -> Dict[str, Any]:
        index = self.build_boond_imputation_index(refresh=refresh)
        projects = index.get("projects") or []
        matched = self.find_best_boond_project_match(project_name, projects)

        if not matched:
            return {
                "input_project_name": project_name,
                "matched_project": None,
                "totals": {
                    "total_days": 0.0,
                    "average_daily_rate": None,
                    "total_cost": None,
                    "cost_status": "missing_rate",
                },
                "by_month": [],
                "meta": index.get("meta", {}),
                "message": "Aucun projet BOOND correspondant trouvé.",
            }

        project_data = next((p for p in projects if clean_text(p.get("project_id")) == matched["project_id"]), None) or {}
        total_days = clean_number(project_data.get("total_days"))
        by_month_map = project_data.get("by_month") or {}

        manual_rate = self.manual_daily_rates_by_project.get(matched["project_id"])
        total_cost = None
        cost_status = "missing_rate"
        if manual_rate is not None:
            total_cost = round(total_days * manual_rate, 2)
            cost_status = "ok"

        by_month = []
        for month, days in sorted((by_month_map or {}).items()):
            days_val = clean_number(days)
            by_month.append({
                "month": month,
                "days": round(days_val, 2),
                "cost": round(days_val * manual_rate, 2) if manual_rate is not None else None,
            })

        return {
            "input_project_name": project_name,
            "matched_project": matched,
            "totals": {
                "total_days": round(total_days, 2),
                "average_daily_rate": manual_rate,
                "total_cost": total_cost,
                "cost_status": cost_status,
            },
            "by_month": by_month,
            "meta": index.get("meta", {}),
            "message": "Coût indisponible: aucun taux journalier fiable trouvé." if manual_rate is None else "OK",
        }

def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text)


def clean_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value).replace("€", "").replace("%", "").replace("\u202f", "").replace(" ", "")
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def safe_int(value: Any) -> int:
    return int(round(clean_number(value)))


def slugify(value: Any) -> str:
    text = clean_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-{2,}", "-", text).strip("-")
    return text or "affaire"


def build_display_name(client: str, affaire: str) -> str:
    parts = [p for p in [clean_text(client), clean_text(affaire)] if p]
    return " - ".join(parts) if parts else "Affaire sans nom"


def month_payload() -> Dict[str, Dict[str, float]]:
    return {m: {"previsionnel": 0.0, "facture": 0.0, "ecart": 0.0} for m in MONTHS}


def row_tuple_to_dict(values: tuple) -> Dict[str, Any]:
    data = {}
    for idx, key in enumerate(COLUMN_ORDER):
        data[key] = values[idx] if idx < len(values) else None
    return data


def anteriorite_from_row(row: Dict[str, Any]) -> float:
    keys = [
        "facturation_cumulee_2017",
        "facturation_cumulee_2018",
        "facturation_cumulee_2021",
        "facturation_cumulee_2022",
        "facturation_cumulee_2023",
        "facturation_cumulee_2024",
        "facturation_cumulee_2025",
    ]
    return sum(clean_number(row.get(k)) for k in keys)


def facture_2026_from_row(row: Dict[str, Any]) -> float:
    return clean_number(row.get("facturation_cumulee_2026"))


class FinanceService:
    def __init__(self, workbook_path: str, sheet_name: str, cache_file: str) -> None:
        self.workbook_path = workbook_path
        self.sheet_name = sheet_name
        self.cache_file = Path(cache_file)
        self._cache: Dict[str, Any] = {
            "status": "idle",
            "generated_at": None,
            "rows_read": 0,
            "rows_kept": 0,
            "affaires_count": 0,
            "warnings": [],
            "items": {},
            "ordered_ids": [],
            "source_path": workbook_path,
            "source_mtime": None,
        }
        self._lock = Lock()

    def workbook_exists(self) -> bool:
        return os.path.exists(self.workbook_path)

    def source_mtime(self) -> Optional[float]:
        try:
            return os.path.getmtime(self.workbook_path)
        except OSError:
            return None

    def cache_status(self) -> Dict[str, Any]:
        cache = self.get_finance_cache()
        return {
            "status": cache.get("status", "unknown"),
            "generated_at": cache.get("generated_at"),
            "rows_read": cache.get("rows_read", 0),
            "rows_kept": cache.get("rows_kept", 0),
            "affaires_count": cache.get("affaires_count", 0),
            "warnings": cache.get("warnings", []),
            "source_path": cache.get("source_path"),
            "source_mtime": cache.get("source_mtime"),
            "cache_file": str(self.cache_file.resolve()),
        }

    def get_finance_cache(self) -> Dict[str, Any]:
        with self._lock:
            cache = dict(self._cache)
        if (
            cache.get("status") == "ready"
            and cache.get("items")
            and cache.get("schema_version") == EXPECTED_SCHEMA_VERSION
        ):
            return cache

        if self.cache_file.exists():
            try:
                disk = json.loads(self.cache_file.read_text(encoding="utf-8"))
                if (
                    disk.get("schema_version") == EXPECTED_SCHEMA_VERSION
                    and disk.get("source_mtime") == self.source_mtime()
                    and disk.get("items")
                ):
                    with self._lock:
                        self._cache = disk
                    return dict(self._cache)
            except Exception:
                logger.exception("Impossible de lire le cache JSON")

        return self.rebuild_finance_cache()

    def rebuild_finance_cache(self) -> Dict[str, Any]:
        if not self.workbook_exists():
            raise FileNotFoundError(f"Fichier introuvable : {self.workbook_path}")

        logger.info("Reconstruction du cache finance depuis %s", self.workbook_path)
        with self._lock:
            self._cache["status"] = "building"

        wb = load_workbook(self.workbook_path, data_only=True, read_only=True)
        if self.sheet_name not in wb.sheetnames:
            raise ValueError(f"Onglet introuvable : {self.sheet_name}")

        ws = wb[self.sheet_name]
        parsed = self.parse_affaires_sheet(ws)

        cache = {
            "schema_version": EXPECTED_SCHEMA_VERSION,
            "status": "ready",
            "generated_at": now_iso(),
            "rows_read": parsed["meta"]["rows_read"],
            "rows_kept": parsed["meta"]["rows_kept"],
            "affaires_count": parsed["meta"]["affaires_count"],
            "warnings": parsed["meta"]["warnings"],
            "source_path": self.workbook_path,
            "source_mtime": self.source_mtime(),
            "sheet_name": self.sheet_name,
            "items": parsed["items"],
            "ordered_ids": parsed["ordered_ids"],
        }
        with self._lock:
            self._cache = cache

        try:
            self.cache_file.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            logger.exception("Impossible d'écrire le cache JSON")

        logger.info(
            "Cache prêt | affaires=%s rows_read=%s rows_kept=%s",
            cache["affaires_count"], cache["rows_read"], cache["rows_kept"],
        )
        return dict(cache)

    def parse_affaires_sheet(self, ws) -> Dict[str, Any]:
        rows_read = 0
        rows_kept = 0
        warnings: List[str] = []
        current_client = ""
        current_parent: Optional[Dict[str, Any]] = None
        blank_streak = 0
        parent_rows: List[Dict[str, Any]] = []

        for row_index, values in enumerate(ws.iter_rows(min_row=13, max_col=len(COLUMN_ORDER), values_only=True), start=13):
            row = row_tuple_to_dict(values)
            raw_client = clean_text(row["client"])
            raw_affaire = clean_text(row["affaire"])
            raw_tag = clean_text(row["tag"])

            if not raw_client and not raw_affaire:
                blank_streak += 1
                if blank_streak >= 6:
                    break
                continue
            blank_streak = 0

            if "pipe operationnel" in slugify(raw_affaire):
                break

            if raw_client:
                current_client = raw_client

            rows_read += 1
            is_parent = bool(raw_client and raw_affaire and not raw_tag)

            if is_parent:
                current_parent = self._normalize_row(row, current_client=current_client)
                current_parent["missions"] = []
                current_parent["_row_index"] = row_index
                parent_rows.append(current_parent)
                rows_kept += 1
                continue

            if not raw_affaire:
                continue

            mission = self._normalize_row(
                row,
                current_client=current_client,
                parent_display=current_parent["display_name"] if current_parent else "",
                parent_id=current_parent["affaire_id"] if current_parent else "",
                force_affaire_name=raw_affaire,
            )

            if current_parent is None:
                mission["missions"] = [self._mission_payload_from_affaire(mission)]
                mission["_row_index"] = row_index
                parent_rows.append(mission)
                current_parent = mission
            else:
                current_parent["missions"].append(self._mission_payload_from_affaire(mission))
            rows_kept += 1

        items: Dict[str, Dict[str, Any]] = {}
        ordered_ids: List[str] = []
        for parent in parent_rows:
            affair = self._finalize_affaire(parent)
            items[affair["affaire_id"]] = affair
            ordered_ids.append(affair["affaire_id"])

        meta = {
            "rows_read": rows_read,
            "rows_kept": rows_kept,
            "affaires_count": len(items),
            "warnings": warnings,
        }
        return {"items": items, "ordered_ids": ordered_ids, "meta": meta}

    def _normalize_row(
        self,
        row: Dict[str, Any],
        current_client: str,
        parent_display: str = "",
        parent_id: str = "",
        force_affaire_name: str = "",
    ) -> Dict[str, Any]:
        client = clean_text(row.get("client")) or current_client
        affaire_name = force_affaire_name or clean_text(row.get("affaire"))
        display_name = build_display_name(client, affaire_name)
        affaire_id = slugify(display_name)

        mensuel = month_payload()
        for month in MONTHS:
            pre = clean_number(row.get(f"{month}_previsionnel"))
            fac = clean_number(row.get(f"{month}_facture"))
            mensuel[month]["previsionnel"] = pre
            mensuel[month]["facture"] = fac
            mensuel[month]["ecart"] = fac - pre

        total_previsionnel = sum(mensuel[m]["previsionnel"] for m in MONTHS)
        total_facture = clean_number(row.get("total_facture")) or sum(mensuel[m]["facture"] for m in MONTHS)
        commande_ht = clean_number(row.get("commande_ht"))
        anteriorite = anteriorite_from_row(row)
        fact_2026 = facture_2026_from_row(row)
        facturation_totale = anteriorite + fact_2026
        reste = clean_number(row.get("reste_a_facturer"))

        return {
            "affaire_id": affaire_id,
            "display_name": display_name,
            "client": client,
            "affaire": affaire_name,
            "numero": clean_text(row.get("numero")),
            "tag": clean_text(row.get("tag")),
            "tags": [clean_text(row.get("tag"))] if clean_text(row.get("tag")) else [],
            "parent_affaire_id": parent_id or affaire_id,
            "parent_display_name": parent_display or display_name,
            "delai_reglement_jours": safe_int(row.get("delai_reglement_jours")),
            "commande_ht": commande_ht,
            "facturation_cumulee_2017": clean_number(row.get("facturation_cumulee_2017")),
            "facturation_cumulee_2018": clean_number(row.get("facturation_cumulee_2018")),
            "facturation_cumulee_2021": clean_number(row.get("facturation_cumulee_2021")),
            "facturation_cumulee_2022": clean_number(row.get("facturation_cumulee_2022")),
            "facturation_cumulee_2023": clean_number(row.get("facturation_cumulee_2023")),
            "facturation_cumulee_2024": clean_number(row.get("facturation_cumulee_2024")),
            "facturation_cumulee_2025": clean_number(row.get("facturation_cumulee_2025")),
            "facturation_cumulee_2026": fact_2026,
            "anteriorite": anteriorite,
            "facture_2026": fact_2026,
            "facturation_totale": facturation_totale,
            "reste_a_facturer": reste,
            "has_reste_value": row.get("reste_a_facturer") not in (None, ""),
            "mensuel": mensuel,
            "total_previsionnel": total_previsionnel,
            "total_facture": total_facture,
            "ecart_previsionnel_vs_facture": total_facture - total_previsionnel,
            "taux_avancement_financier": (facturation_totale / commande_ht) if commande_ht else 0.0,
        }

    def _mission_payload_from_affaire(self, data: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "tag": data.get("tag", ""),
            "label": data.get("affaire", ""),
            "numero": data.get("numero", ""),
            "commande_ht": data.get("commande_ht", 0.0),
            "facturation_cumulee_2017": data.get("facturation_cumulee_2017", 0.0),
            "facturation_cumulee_2018": data.get("facturation_cumulee_2018", 0.0),
            "facturation_cumulee_2021": data.get("facturation_cumulee_2021", 0.0),
            "facturation_cumulee_2022": data.get("facturation_cumulee_2022", 0.0),
            "facturation_cumulee_2023": data.get("facturation_cumulee_2023", 0.0),
            "facturation_cumulee_2024": data.get("facturation_cumulee_2024", 0.0),
            "facturation_cumulee_2025": data.get("facturation_cumulee_2025", 0.0),
            "facturation_cumulee_2026": data.get("facturation_cumulee_2026", 0.0),
            "anteriorite": data.get("anteriorite", 0.0),
            "facture_2026": data.get("facture_2026", data.get("facturation_cumulee_2026", 0.0)),
            "facturation_totale": (
                clean_number(data.get("anteriorite", 0.0))
                + clean_number(data.get("facture_2026", data.get("facturation_cumulee_2026", 0.0)))
            ),
            "reste_a_facturer": data.get("reste_a_facturer", 0.0),
            "has_reste_value": data.get("has_reste_value", False),
            "total_previsionnel": data.get("total_previsionnel", 0.0),
            "total_facture": data.get("total_facture", 0.0),
            "mensuel": data.get("mensuel", month_payload()),
        }

    def _finalize_affaire(self, affaire: Dict[str, Any]) -> Dict[str, Any]:
        excel_anteriorite = anteriorite_from_row(affaire)
        excel_facture_2026 = clean_number(affaire.get("facturation_cumulee_2026"))
        excel_reste_a_facturer = clean_number(affaire.get("reste_a_facturer"))

        missions = affaire.get("missions") or []
        if missions:
            monthly = month_payload()
            tags = []
            numero = affaire.get("numero", "")
            commande = 0.0
            facture_2026 = 0.0
            anteriorite = 0.0
            delai_values = []

            for mission in missions:
                commande += clean_number(mission.get("commande_ht"))
                anteriorite += clean_number(mission.get("anteriorite"))
                facture_2026 += clean_number(mission.get("facture_2026", mission.get("facturation_cumulee_2026")))
                if clean_number(mission.get("delai_reglement_jours")) > 0:
                    delai_values.append(safe_int(mission.get("delai_reglement_jours")))
                if mission.get("tag"):
                    tags.append(clean_text(mission["tag"]))
                if not numero and mission.get("numero"):
                    numero = clean_text(mission["numero"])

                mm = mission.get("mensuel") or {}
                for month in MONTHS:
                    monthly[month]["previsionnel"] += clean_number((mm.get(month) or {}).get("previsionnel"))
                    monthly[month]["facture"] += clean_number((mm.get(month) or {}).get("facture"))

            for month in MONTHS:
                monthly[month]["ecart"] = monthly[month]["facture"] - monthly[month]["previsionnel"]

            affaire["commande_ht"] = commande
            affaire["anteriorite"] = anteriorite
            affaire["facture_2026"] = facture_2026
            affaire["facturation_totale"] = anteriorite + facture_2026
            affaire["reste_a_facturer"] = sum(clean_number(m.get("reste_a_facturer")) for m in missions)
            affaire["facturation_cumulee_2026"] = affaire["facture_2026"]
            affaire["mensuel"] = monthly
            affaire["tags"] = sorted(set(t for t in tags if t))
            affaire["numero"] = numero
            affaire["total_previsionnel"] = sum(monthly[m]["previsionnel"] for m in MONTHS)
            affaire["total_facture"] = sum(monthly[m]["facture"] for m in MONTHS)
            affaire["ecart_previsionnel_vs_facture"] = affaire["total_facture"] - affaire["total_previsionnel"]
            affaire["delai_reglement_jours"] = affaire.get("delai_reglement_jours") or (max(delai_values) if delai_values else 0)
            affaire["taux_avancement_financier"] = (affaire["facturation_totale"] / affaire["commande_ht"]) if affaire["commande_ht"] else 0.0
        else:
            affaire["missions"] = []
            if affaire.get("tag") and not affaire.get("tags"):
                affaire["tags"] = [affaire["tag"]]
            affaire["anteriorite"] = excel_anteriorite
            affaire["facture_2026"] = excel_facture_2026
            affaire["facturation_totale"] = affaire["anteriorite"] + affaire["facture_2026"]
            if abs(clean_number(affaire.get("reste_a_facturer"))) < 1e-9 and abs(clean_number(affaire.get("commande_ht"))) > 1e-9:
                affaire["reste_a_facturer"] = clean_number(affaire.get("commande_ht")) - affaire["facturation_totale"]
            affaire["taux_avancement_financier"] = (affaire["facturation_totale"] / clean_number(affaire.get("commande_ht"))) if clean_number(affaire.get("commande_ht")) else 0.0

        affaire["audit"] = {
            "excel_anteriorite": excel_anteriorite,
            "excel_facture_2026": excel_facture_2026,
            "excel_reste_a_facturer": excel_reste_a_facturer,
            "missions_anteriorite": clean_number(affaire.get("anteriorite", 0)),
            "missions_facture_2026": clean_number(affaire.get("facture_2026", 0)),
            "missions_reste_a_facturer": clean_number(affaire.get("reste_a_facturer", 0)),
            "ecart_anteriorite": clean_number(affaire.get("anteriorite", 0)) - excel_anteriorite,
            "ecart_facture_2026": clean_number(affaire.get("facture_2026", 0)) - excel_facture_2026,
            "ecart_reste_a_facturer": clean_number(affaire.get("reste_a_facturer", 0)) - excel_reste_a_facturer,
        }
        affaire["insights"] = self.compute_insights(affaire)
        affaire.pop("has_reste_value", None)
        affaire.pop("tag", None)
        affaire.pop("_row_index", None)
        return affaire

    @staticmethod
    def compute_insights(affaire: Dict[str, Any]) -> List[str]:
        insights: List[str] = []
        commande = clean_number(affaire.get("commande_ht"))
        facture = clean_number(affaire.get("facturation_totale"))
        reste = clean_number(affaire.get("reste_a_facturer"))
        prev = clean_number(affaire.get("total_previsionnel"))
        fact_total = clean_number(affaire.get("total_facture"))
        taux = float(affaire.get("taux_avancement_financier") or 0.0)

        if commande > 0 and facture <= 0:
            insights.append("Aucune facturation enregistrée alors qu'une commande existe.")

        if reste > max(1000.0, commande * 0.4):
            insights.append("Le reste à facturer est élevé.")

        if prev > 0 and fact_total < prev * 0.75:
            insights.append("La facturation est en retard par rapport au prévisionnel.")
        elif prev > 0 and fact_total > prev * 1.1:
            insights.append("La facturation est en avance par rapport au prévisionnel.")

        if commande > 0 and taux >= 0.9:
            insights.append("L'affaire est presque finalisée financièrement.")

        pointage_amount = clean_number(affaire.get("pointage_progress_amount"))
        if commande > 0 and pointage_amount > 0:
            gap = facture - pointage_amount
            if abs(gap) > max(1000.0, commande * 0.1):
                direction = "en avance" if gap > 0 else "en retard"
                insights.append(f"La facturation cumulée est {direction} de {abs(gap):,.0f} € vs avancement pointé CST.".replace(',', ' '))

        if not insights:
            insights.append("Situation financière stable selon les données disponibles.")
        return insights

    @staticmethod
    def lightweight_affaires(cache: Dict[str, Any], search: str = "") -> List[Dict[str, str]]:
        items = cache.get("items", {}) or {}
        q = slugify(search) if search else ""
        out = []
        for affaire_id, item in items.items():
            display = item.get("display_name") or affaire_id
            if q and q not in slugify(display):
                continue
            out.append({"affaire_id": affaire_id, "display_name": display})
        out.sort(key=lambda x: x["display_name"])
        return out



class MetronomeService:
    _PROJECT_MATCH_IGNORED_TOKENS = {
        "de", "du", "des", "la", "le", "les", "a", "au", "aux", "et",
        "reunion", "syt", "projet", "affaire",
    }

    def __init__(self, base_path: str) -> None:
        self.base_path = Path(base_path)
        self._lock = Lock()
        self._cache: Dict[str, Any] = {"loaded": False, "tables": {}, "mtime": {}}

    def _read_csv_rows(self, path: Path) -> List[Dict[str, str]]:
        raw = path.read_bytes()
        for enc in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                text = raw.decode(enc)
                break
            except Exception:
                continue
        else:
            text = raw.decode("utf-8", errors="ignore")
        return list(csv.DictReader(io.StringIO(text), delimiter=","))

    def _table_path(self, key: str) -> Path:
        return self.base_path / METRONOME_FILES[key]

    def _current_mtime(self) -> Dict[str, Optional[float]]:
        out: Dict[str, Optional[float]] = {}
        for key in METRONOME_FILES:
            p = self._table_path(key)
            try:
                out[key] = p.stat().st_mtime
            except OSError:
                out[key] = None
        return out

    def _ensure_loaded(self) -> Dict[str, Any]:
        with self._lock:
            mt = self._current_mtime()
            if self._cache.get("loaded") and self._cache.get("mtime") == mt:
                return dict(self._cache)

            tables: Dict[str, List[Dict[str, str]]] = {}
            missing = []
            for key in METRONOME_FILES:
                p = self._table_path(key)
                if not p.exists():
                    missing.append(str(p))
                    tables[key] = []
                    continue
                tables[key] = self._read_csv_rows(p)

            self._cache = {
                "loaded": True,
                "tables": tables,
                "mtime": mt,
                "missing": missing,
                "loaded_at": now_iso(),
            }
            return dict(self._cache)

    @staticmethod
    def _row_id(row: Dict[str, str]) -> str:
        return clean_text(row.get("🔒 Row ID") or row.get("Row ID") or row.get("ID"))

    @staticmethod
    def _idx(rows: List[Dict[str, str]], key: str = "ID") -> Dict[str, Dict[str, str]]:
        out: Dict[str, Dict[str, str]] = {}
        for r in rows:
            k = clean_text(r.get(key)) or MetronomeService._row_id(r)
            if k:
                out[k] = r
        return out

    def _tokenize_project_name(self, value: str) -> set[str]:
        return {
            token for token in slugify(value).split("-")
            if token and token not in self._PROJECT_MATCH_IGNORED_TOKENS and len(token) > 1
        }

    @staticmethod
    def _get_first_value(row: Dict[str, str], keys: List[str]) -> str:
        for key in keys:
            val = clean_text(row.get(key))
            if val:
                return val
        return ""
    @staticmethod
    def _is_empty_value(value: Any) -> bool:
        txt = clean_text(value).lower()
        return txt in {"", "none", "null", "nan", "na"}

    def _split_multi_values(self, value: Any) -> List[str]:
        if value is None:
            return []
        if isinstance(value, list):
            parts = value
        else:
            parts = str(value).split(",")
        out: List[str] = []
        for p in parts:
            t = clean_text(p)
            if self._is_empty_value(t):
                continue
            out.append(t)
        return out

    @staticmethod
    def _parse_bool(value: Any) -> bool:
        txt = clean_text(value).lower()
        return txt in {"true", "1", "yes", "oui", "vrai"}

    @staticmethod
    def _parse_date_value(value: str) -> Optional[datetime]:
        txt = clean_text(value)
        if not txt:
            return None
        # Nettoie les préfixes type jour abrégé: "Mar 09/12/25" -> "09/12/25"
        txt = re.sub(r"^(?:lun|mar|mer|jeu|ven|sam|dim|mon|tue|wed|thu|fri|sat|sun)\.?\s+", "", txt, flags=re.IGNORECASE)
        for fmt in (
            "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%m/%d/%Y", "%m/%d/%y",
            "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y %H:%M:%S", "%m/%d/%y %H:%M",
            "%m/%d/%Y %I:%M %p", "%m/%d/%y %I:%M %p", "%d/%m/%y %H:%M:%S"
        ):
            for candidate in (txt, txt[:19]):
                try:
                    return datetime.strptime(candidate, fmt)
                except Exception:
                    continue
        simple = slugify(txt).replace("-", " ")
        match = re.match(r"^(\d{1,2})\s+([a-z]+)\s+(\d{4})$", simple)
        if match:
            day = int(match.group(1))
            month_txt = match.group(2)
            year = int(match.group(3))
            months = {
                "janvier": 1, "fevrier": 2, "mars": 3, "avril": 4, "mai": 5, "juin": 6,
                "juillet": 7, "aout": 8, "septembre": 9, "octobre": 10, "novembre": 11, "decembre": 12,
            }
            month = months.get(month_txt)
            if month:
                try:
                    return datetime(year, month, day)
                except Exception:
                    return None
        return None

    @classmethod
    def _parse_date_only(cls, value: str) -> Optional[date]:
        dt = cls._parse_date_value(value)
        return dt.date() if dt else None

    @staticmethod
    def _normalize_company(value: Any) -> str:
        return clean_text(value) or "Non renseigné"

    @staticmethod
    def _normalize_zone(value: Any) -> str:
        return clean_text(value) or "Général"

    @staticmethod
    def _normalize_lot(value: Any) -> str:
        return clean_text(value) or "Sans lot"

    @staticmethod
    def _normalize_owner(value: Any) -> str:
        return clean_text(value) or "Non attribué"

    @staticmethod
    def _company_logo_from_row(row: Dict[str, Any]) -> str:
        for key in ["Logo", "Logo URL", "Avatar", "Avatar URL", "Icon", "Icon URL"]:
            val = clean_text(row.get(key))
            if val:
                return val
        return ""

    @staticmethod
    def reminder_level(deadline: Optional[date], completed: bool, ref_date: date) -> Optional[int]:
        if completed or not deadline:
            return None
        days_late = (ref_date - deadline).days
        if days_late <= 0:
            return None
        return ((days_late - 1) // 7) + 1

    @staticmethod
    def _business_day_delta(target: date, ref_date: date) -> int:
        if target == ref_date:
            return 0
        step = 1 if target > ref_date else -1
        cur = ref_date
        count = 0
        while cur != target:
            cur = cur + timedelta(days=step)
            if cur.weekday() < 5:
                count += step
        return count

    def reminders_by_company(self, rem_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        if not rem_rows:
            return []
        grouped: Dict[str, int] = {}
        logos: Dict[str, str] = {}
        for row in rem_rows:
            company = self._normalize_company(row.get("__company__"))
            grouped[company] = grouped.get(company, 0) + 1
            logos[company] = clean_text(row.get("__logo__"))
        return [
            {"name": name, "count": count, "logo": logos.get(name, "")}
            for name, count in sorted(grouped.items(), key=lambda x: (-x[1], x[0]))
        ]

    def reminders_for_project(
        self,
        project_title: str,
        ref_date: date,
        max_level: int = 8,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        entries_override: Optional[List[Dict[str, Any]]] = None,
        company_logos: Optional[Dict[str, str]] = None,
    ) -> List[Dict[str, Any]]:
        source = entries_override or []
        rows: List[Dict[str, Any]] = []
        for e in source:
            if not e.get("is_task"):
                continue
            created_date = self._parse_date_only(e.get("meeting_date", ""))
            if start_date and created_date and created_date < start_date:
                continue
            if end_date and created_date and created_date > end_date:
                continue
            completed = bool(e.get("is_closed"))
            done = e.get("done_date")
            if done is not None:
                completed = True
            deadline = self._parse_date_only(e.get("deadline", ""))
            lvl = self.reminder_level(deadline, completed, ref_date)
            if lvl is None:
                continue
            if int(lvl) > max_level:
                continue
            company = self._normalize_company(e.get("company"))
            zones = e.get("area_names") or ["Général"]
            if not zones:
                zones = ["Général"]
            for zone in zones:
                rows.append({
                    **e,
                    "__project__": project_title,
                    "__deadline__": deadline,
                    "__reminder__": int(lvl),
                    "__zone__": self._normalize_zone(zone),
                    "__company__": company,
                    "__logo__": clean_text((company_logos or {}).get(company, "")),
                })
        rows.sort(key=lambda r: (-r["__reminder__"], r.get("__deadline__") or date.max))
        return rows

    def followups_for_project(
        self,
        project_title: str,
        ref_date: date,
        exclude_entry_ids: Optional[List[str]],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        entries_override: Optional[List[Dict[str, Any]]] = None,
        company_logos: Optional[Dict[str, str]] = None,
    ) -> List[Dict[str, Any]]:
        source = entries_override or []
        excluded = {clean_text(v) for v in (exclude_entry_ids or []) if clean_text(v)}
        rows: List[Dict[str, Any]] = []
        for e in source:
            if not e.get("is_task"):
                continue
            eid = clean_text(e.get("entry_id"))
            if eid and eid in excluded:
                continue
            created_date = self._parse_date_only(e.get("meeting_date", ""))
            if start_date and created_date and created_date < start_date:
                continue
            if end_date and created_date and created_date > end_date:
                continue
            completed = bool(e.get("is_closed"))
            done = e.get("done_date")
            if done is not None:
                completed = True
            if completed:
                continue
            deadline = self._parse_date_only(e.get("deadline", ""))
            if deadline and deadline < ref_date:
                continue
            company = self._normalize_company(e.get("company"))
            zones = e.get("area_names") or ["Général"]
            if not zones:
                zones = ["Général"]
            for zone in zones:
                rows.append({
                    **e,
                    "__id__": eid,
                    "__project__": project_title,
                    "__deadline__": deadline,
                    "__zone__": self._normalize_zone(zone),
                    "__company__": company,
                    "__deadline_sort__": deadline or date.max,
                })
        rows.sort(key=lambda r: (r.get("__deadline_sort__") or date.max, r.get("__company__") or ""))
        return rows

    def meeting_simple_kpis(self, entries_rows: List[Dict[str, Any]], ref_date: date) -> Dict[str, int]:
        tasks = [e for e in entries_rows if e.get("is_task")]
        memos = [e for e in entries_rows if e.get("is_memo")]
        open_tasks = [e for e in tasks if not bool(e.get("is_closed"))]
        closed_tasks = [e for e in tasks if bool(e.get("is_closed"))]
        late_tasks = 0
        for e in open_tasks:
            dl = self._parse_date_only(e.get("deadline", ""))
            if dl and dl < ref_date:
                late_tasks += 1
        return {
            "total_entries": len(entries_rows),
            "tasks_meeting": len(tasks),
            "memos_meeting": len(memos),
            "open_tasks": len(open_tasks),
            "closed_tasks": len(closed_tasks),
            "late_tasks": late_tasks,
        }


    def _score_project_match(self, target_slug: str, target_tokens: set[str], candidate_name: str) -> int:
        candidate_slug = slugify(candidate_name)
        if not candidate_slug:
            return 0
        score = 0
        if candidate_slug == target_slug:
            score = 100
        else:
            if target_slug and target_slug in candidate_slug:
                score = max(score, 80)
            if candidate_slug and candidate_slug in target_slug:
                score = max(score, 78)
            candidate_tokens = self._tokenize_project_name(candidate_name)
            common_tokens = target_tokens & candidate_tokens
            if common_tokens:
                token_score = 45 + len(common_tokens) * 12
                if target_tokens and candidate_tokens and common_tokens == target_tokens == candidate_tokens:
                    token_score += 10
                score = max(score, min(token_score, 95))
        return score

    def _resolve_project(self, projects: List[Dict[str, str]], entries: List[Dict[str, str]], project_name: str) -> Dict[str, Any]:
        target_name = clean_text(project_name)
        target_slug = slugify(target_name)
        target_tokens = self._tokenize_project_name(target_name)
        resolved: Dict[str, Any] = {
            "searched_project_name": target_name,
            "searched_project_slug": target_slug,
            "resolved_project_title": "",
            "resolved_project_id": "",
            "matched_project_name": "",
            "matched_project_slug": "",
            "match_score": 0,
            "resolution_mode": "fallback",
            "best_project_candidates": [],
            "best_entry_candidates": [],
        }
        if not target_slug:
            return resolved

        scored_projects: List[Dict[str, Any]] = []
        best_project: Optional[Dict[str, str]] = None
        best_name = ""
        best_score = 0
        best_len = 10**9

        for project in projects:
            title = self._get_first_value(project, METRONOME_COLUMN_ALIASES["project_title_projects"])
            name = self._get_first_value(project, METRONOME_COLUMN_ALIASES["project_name_projects"])
            project_row_id = self._row_id(project)
            for candidate_name in [title, name]:
                if not candidate_name:
                    continue
                score = self._score_project_match(target_slug, target_tokens, candidate_name)
                if score <= 0:
                    continue
                scored_projects.append({
                    "project_id": project_row_id,
                    "project_name": candidate_name,
                    "project_slug": slugify(candidate_name),
                    "score": score,
                })
                slug_len_delta = abs(len(slugify(candidate_name)) - len(target_slug))
                if score > best_score or (score == best_score and slug_len_delta < best_len):
                    best_project = project
                    best_name = candidate_name
                    best_score = score
                    best_len = slug_len_delta

        scored_projects.sort(key=lambda x: (-x["score"], abs(len(x["project_slug"]) - len(target_slug))))
        resolved["best_project_candidates"] = scored_projects[:5]

        min_reliable_score = 55
        if best_project and best_score >= min_reliable_score:
            resolved_title = self._get_first_value(best_project, METRONOME_COLUMN_ALIASES["project_title_projects"])
            resolved_id = self._row_id(best_project)
            mode = "title" if clean_text(best_project.get("Title")) else "name"
            if resolved_id:
                mode = "id"
            resolved.update({
                "resolved_project_title": resolved_title,
                "resolved_project_id": resolved_id,
                "matched_project_name": best_name,
                "matched_project_slug": slugify(best_name),
                "match_score": best_score,
                "resolution_mode": mode,
            })
            return resolved

        scored_entry_titles: Dict[str, Dict[str, Any]] = {}
        for e in entries:
            entry_title = self._get_first_value(e, METRONOME_COLUMN_ALIASES["project_title_entries"])
            if not entry_title:
                continue
            score = self._score_project_match(target_slug, target_tokens, entry_title)
            if score <= 0:
                continue
            eid = self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_project_id"])
            existing = scored_entry_titles.get(entry_title)
            if not existing or score > existing["score"]:
                scored_entry_titles[entry_title] = {
                    "project_id": eid,
                    "project_name": entry_title,
                    "project_slug": slugify(entry_title),
                    "score": score,
                }

        scored_entries = sorted(scored_entry_titles.values(), key=lambda x: (-x["score"], abs(len(x["project_slug"]) - len(target_slug))))
        resolved["best_entry_candidates"] = scored_entries[:5]

        if scored_entries and scored_entries[0]["score"] >= min_reliable_score:
            best = scored_entries[0]
            resolved.update({
                "resolved_project_title": best["project_name"],
                "resolved_project_id": clean_text(best.get("project_id")),
                "matched_project_name": best["project_name"],
                "matched_project_slug": best["project_slug"],
                "match_score": best["score"],
                "resolution_mode": "entries_fallback",
            })
        return resolved

    def build_project_board(self, project_name: str, start_date: str = "", end_date: str = "") -> Dict[str, Any]:
        cache = self._ensure_loaded()
        t = cache.get("tables", {})
        projects = t.get("projects", [])
        entries = t.get("entries", [])
        meetings_rows = t.get("meetings", [])
        areas_rows = t.get("areas", [])
        packages_rows = t.get("packages", [])
        companies_rows = t.get("companies", [])
        users_rows = t.get("users", [])
        comments_rows = t.get("comments", [])
        documents_rows = t.get("documents", [])

        meetings = self._idx(meetings_rows)
        areas = self._idx(areas_rows)
        packages = self._idx(packages_rows)
        companies = self._idx(companies_rows)
        users = self._idx(users_rows)
        company_logos_by_name: Dict[str, str] = {}
        for comp in companies_rows:
            cname = self._normalize_company(comp.get("Name"))
            logo = self._company_logo_from_row(comp)
            if cname not in company_logos_by_name or logo:
                company_logos_by_name[cname] = logo

        comments_by_entry: Dict[str, List[str]] = {}
        memo_comments_by_entry: Dict[str, List[Dict[str, Any]]] = {}
        for c in comments_rows:
            eid = self._get_first_value(c, METRONOME_COLUMN_ALIASES["comment_memo_id"])
            txt = clean_text(c.get("Comment") or c.get("Text") or c.get("Entry"))
            owner_id = self._get_first_value(c, METRONOME_COLUMN_ALIASES["comment_owner_id"])
            created_at = self._parse_date_value(self._get_first_value(c, METRONOME_COLUMN_ALIASES["comment_date"]))
            if eid and txt:
                comments_by_entry.setdefault(eid, []).append(txt)
                memo_comments_by_entry.setdefault(eid, []).append({
                    "comment_id": self._row_id(c),
                    "memo_id": eid,
                    "text": txt,
                    "owner_user_id": owner_id,
                    "owner_user_name": clean_text(users.get(owner_id, {}).get("Name")),
                    "created_at": created_at.isoformat(sep=" ") if created_at else "",
                    "raw": c,
                })

        target = clean_text(project_name)
        match_debug = self._resolve_project(projects, entries, target)
        resolved_title = clean_text(match_debug.get("resolved_project_title"))
        resolved_id = clean_text(match_debug.get("resolved_project_id"))

        project_info: Dict[str, str] = {}
        if resolved_title:
            for p in projects:
                pt = self._get_first_value(p, METRONOME_COLUMN_ALIASES["project_title_projects"])
                if clean_text(pt) == resolved_title:
                    project_info = p
                    if not resolved_id:
                        resolved_id = self._row_id(p)
                        match_debug["resolved_project_id"] = resolved_id
                    break
        if not project_info and resolved_id:
            for p in projects:
                if self._row_id(p) == resolved_id:
                    project_info = p
                    if not resolved_title:
                        resolved_title = self._get_first_value(p, METRONOME_COLUMN_ALIASES["project_title_projects"])
                        match_debug["resolved_project_title"] = resolved_title
                    break

        if not resolved_title and not resolved_id:
            return {
                "ok": False,
                "project_name": target,
                "reason": "project_not_found",
                "confidence_level": "low",
                "warning": True,
                "warning_message": "Aucune correspondance projet fiable trouvée",
                "match_debug": match_debug,
                "missing_files": cache.get("missing", []),
                "loaded_at": cache.get("loaded_at"),
            }

        rows = []
        today = datetime.now().date()
        rows_filtered_by_title = 0
        rows_filtered_by_id = 0
        filter_mode = "id" if resolved_id else "title"
        match_debug["filter_mode"] = filter_mode

        def parse_date(value: str) -> Optional[date]:
            txt = clean_text(value)
            if not txt:
                return None
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(txt[:10], fmt).date()
                except Exception:
                    continue
            return self._parse_date_only(txt)

        selected_entries: List[Dict[str, Any]] = []

        for e in entries:
            entry_project_title = self._get_first_value(e, METRONOME_COLUMN_ALIASES["project_title_entries"])
            entry_project_id = self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_project_id"])

            use_row = False
            if resolved_id:
                if entry_project_id == resolved_id:
                    rows_filtered_by_id += 1
                    use_row = True
            elif resolved_title and entry_project_title == resolved_title:
                rows_filtered_by_title += 1
                use_row = True

            if not use_row:
                continue

            completed_flag = self._parse_bool(self._get_first_value(e, ["Completed/true/false", "Completed"]))
            done_date = self._parse_date_only(self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_done_date"]))
            is_closed = bool(completed_flag)
            if done_date is not None:
                is_closed = True

            entry_id = self._row_id(e)
            area_ids = self._split_multi_values(self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_area_ids"]))
            area = areas.get(area_ids[0], {}) if area_ids else {}
            task_pkg_id = self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_task_package_id"])
            pkg = packages.get(task_pkg_id, {})
            company_id_for_task = self._get_first_value(e, ["Company/ID"])
            company_name_for_task = self._get_first_value(e, ["Company/Name for Tasks", "Company"])
            pkg_company_id = self._get_first_value(pkg, METRONOME_COLUMN_ALIASES["package_company_id"])
            company = companies.get(company_id_for_task) or companies.get(pkg_company_id) or {}
            owner_user_id = self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_owner_id"])
            user = users.get(owner_user_id, {})
            meeting_id = self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_meeting_id"])
            meeting = meetings.get(meeting_id, {})
            due = self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_deadline"])
            meeting_date = self._get_first_value(meeting, METRONOME_COLUMN_ALIASES["meeting_date"])
            request_date = parse_date(due) or parse_date(meeting_date)

            category_label = self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_category"]) or ""
            category_slug = slugify(category_label)
            is_task_entry = "tache" in category_slug
            is_memo_entry = "memo" in category_slug
            memo_package_ids = self._split_multi_values(self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_memo_package_ids"]))
            memo_package_labels = [clean_text(packages.get(pid, {}).get("Name")) for pid in memo_package_ids if pid]
            created_by_id = self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_created_by_id"])
            last_comment_text = clean_text(self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_comment_text"]))
            memo_comments = memo_comments_by_entry.get(entry_id, []) if is_memo_entry else []
            selected_entries.append({
                "entry_id": entry_id,
                "entry_title": clean_text(e.get("Title")),
                "entry_type": "Tâche" if is_task_entry else ("Mémo" if is_memo_entry else category_label),
                "is_task": is_task_entry,
                "is_memo": is_memo_entry,
                "project_id": resolved_id,
                "project_title": resolved_title,
                "project_archived": self._parse_bool(self._get_first_value(project_info, ["Archived", "Is Archived"])),
                "meeting_id": meeting_id,
                "meeting_date": meeting_date,
                "created_by_user_id": created_by_id,
                "created_by_user_name": clean_text(users.get(created_by_id, {}).get("Name")),
                "package_id": task_pkg_id,
                "package_name": clean_text(pkg.get("Name")),
                "package_label": clean_text(pkg.get("Label") or pkg.get("Name")),
                "package_company_id": pkg_company_id,
                "package_company_name": clean_text(companies.get(pkg_company_id, {}).get("Name")),
                "memo_package_ids": memo_package_ids,
                "memo_package_labels": [x for x in memo_package_labels if x],
                "area_ids": area_ids,
                "area_names": [clean_text(areas.get(aid, {}).get("Name")) for aid in area_ids if aid],
                "owner_user_id": owner_user_id,
                "owner_full_name": clean_text(user.get("Name")),
                "deadline": due,
                "status_id": self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_status_id"]),
                "status_label": self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_status_label"]),
                "completed": is_closed,
                "completed_state": "closed" if is_closed else "open",
                "completed_at": done_date.isoformat() if done_date else "",
                "first_response_at": memo_comments[0]["created_at"] if memo_comments else "",
                "last_activity_at": (memo_comments[-1]["created_at"] if memo_comments else (done_date.isoformat() if done_date else "")),
                "last_comment_text": last_comment_text or (memo_comments[-1]["text"] if memo_comments else ""),
                "last_comment_date": memo_comments[-1]["created_at"] if memo_comments else "",
                "last_comment_user_id": memo_comments[-1]["owner_user_id"] if memo_comments else self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_comment_editor_id"]),
                "last_comment_user_name": memo_comments[-1]["owner_user_name"] if memo_comments else clean_text(users.get(self._get_first_value(e, METRONOME_COLUMN_ALIASES["entry_comment_editor_id"]), {}).get("Name")),
                "memo_comments_count": len(memo_comments),
                "memo_comments": memo_comments,
                "request_date": request_date,
                "done_date": done_date,
                "is_closed": is_closed,
                "company": clean_text(company_name_for_task) or clean_text(company.get("Name")) or clean_text(companies.get(pkg_company_id, {}).get("Name")) or "Non renseigné",
                "raw": e,
            })

            if is_closed or not is_task_entry:
                continue

            overdue = False
            if due:
                due_date = parse_date(due)
                overdue = bool(due_date and due_date < today)

            rows.append({
                "zone": self._normalize_zone(area.get("Name")),
                "lot": self._normalize_lot(pkg.get("Name")),
                "sujet": clean_text(e.get("Title")),
                "entreprise": self._normalize_company(clean_text(company_name_for_task) or clean_text(company.get("Name"))),
                "responsable": self._normalize_owner(user.get("Name")),
                "statut": clean_text(e.get("Status")),
                "date_echeance": due,
                "reunion_origine": meeting_date,
                "commentaire": " | ".join(comments_by_entry.get(entry_id, [])),
                "overdue": overdue,
            })

        match_debug["rows_filtered_by_title"] = rows_filtered_by_title
        match_debug["rows_filtered_by_id"] = rows_filtered_by_id

        def count_by(field: str) -> List[Dict[str, Any]]:
            agg: Dict[str, int] = {}
            for r in rows:
                k = clean_text(r.get(field)) or "Non renseigné"
                agg[k] = agg.get(k, 0) + 1
            return [{"label": k, "count": v} for k, v in sorted(agg.items(), key=lambda x: (-x[1], x[0]))]

        by_meeting = count_by("reunion_origine")
        by_company = count_by("entreprise")

        range_start = parse_date(start_date) if start_date else None
        range_end = parse_date(end_date) if end_date else None
        meeting_dates = [parse_date(e.get("meeting_date", "")) for e in selected_entries if parse_date(e.get("meeting_date", ""))]
        meeting_ref_date = max(meeting_dates) if meeting_dates else None
        reference_date = range_end or meeting_ref_date or today
        reference_date_text = reference_date.strftime("%d/%m/%Y")

        current_meeting_entry_ids: List[str] = []
        if meeting_ref_date:
            current_meeting_entry_ids = [
                clean_text(e.get("entry_id"))
                for e in selected_entries
                if parse_date(e.get("meeting_date", "")) == meeting_ref_date and clean_text(e.get("entry_id"))
            ]

        rem_rows = self.reminders_for_project(
            resolved_title or target,
            reference_date,
            max_level=8,
            start_date=range_start,
            end_date=range_end,
            entries_override=selected_entries,
            company_logos=company_logos_by_name,
        )
        fol_rows = self.followups_for_project(
            resolved_title or target,
            reference_date,
            exclude_entry_ids=current_meeting_entry_ids,
            start_date=range_start,
            end_date=range_end,
            entries_override=selected_entries,
        )
        due_by_company_list = self.reminders_by_company(rem_rows)
        meeting_simple_kpis = self.meeting_simple_kpis(selected_entries, reference_date)

        reminder_threshold_days = 7
        open_stats: Dict[str, Dict[str, int]] = {}
        delay_stats: Dict[str, Dict[str, int]] = {}
        for item in selected_entries:
            company = item["company"]
            request_date = item.get("request_date")
            done_date = item.get("done_date")
            if not item.get("is_closed"):
                row = open_stats.setdefault(company, {"open_count": 0, "reminder_count": 0})
                row["open_count"] += 1
                if request_date and (today - request_date).days >= reminder_threshold_days:
                    row["reminder_count"] += 1
            elif request_date and done_date:
                delay_days = max(0, (done_date - request_date).days)
                row = delay_stats.setdefault(company, {"sum_days": 0, "count": 0})
                row["sum_days"] += delay_days
                row["count"] += 1

        open_tasks_by_company = [
            {"label": k, "open_count": v["open_count"], "reminder_count": v["reminder_count"]}
            for k, v in sorted(open_stats.items(), key=lambda x: (-x[1]["open_count"], -x[1]["reminder_count"], x[0]))
        ]
        average_processing_days_by_company = [
            {"label": k, "avg_days": round(v["sum_days"] / max(1, v["count"]), 1), "closed_count": v["count"]}
            for k, v in sorted(delay_stats.items(), key=lambda x: (x[1]["sum_days"] / max(1, x[1]["count"]), x[0]))
        ]

        fact_tasks: List[Dict[str, Any]] = []
        fact_memos: List[Dict[str, Any]] = []
        fact_memo_comments: List[Dict[str, Any]] = []
        fact_interactions: List[Dict[str, Any]] = []

        for item in selected_entries:
            deadline_dt = parse_date(item.get("deadline", ""))
            request_dt = item.get("request_date")
            done_dt = item.get("done_date")
            first_response_dt = self._parse_date_value(item.get("first_response_at", "")) if item.get("first_response_at") else None
            last_activity_dt = self._parse_date_value(item.get("last_activity_at", "")) if item.get("last_activity_at") else None

            if item.get("is_task"):
                is_open = not item.get("is_closed")
                is_closed = bool(item.get("is_closed"))
                is_late = bool(is_open and deadline_dt and deadline_dt < today)
                is_closed_late = bool(is_closed and deadline_dt and done_dt and done_dt > deadline_dt)
                is_closed_on_time = bool(is_closed and deadline_dt and done_dt and done_dt <= deadline_dt)
                days_late = max(0, (today - deadline_dt).days) if is_late and deadline_dt else 0
                age_days = max(0, (today - request_dt).days) if is_open and request_dt else 0
                days_to_deadline = (deadline_dt - today).days if deadline_dt else None
                response_delay_days = (first_response_dt.date() - request_dt).days if (first_response_dt and request_dt) else None

                fact_tasks.append({
                    **item,
                    "is_open": is_open,
                    "is_closed": is_closed,
                    "is_late": is_late,
                    "is_closed_late": is_closed_late,
                    "is_closed_on_time": is_closed_on_time,
                    "days_late": days_late,
                    "age_days": age_days,
                    "days_to_deadline": days_to_deadline,
                    "response_delay_days": response_delay_days,
                    "completed_at": item.get("completed_at") or (done_dt.isoformat() if done_dt else ""),
                    "first_response_at": item.get("first_response_at") or (first_response_dt.isoformat(sep=" ") if first_response_dt else ""),
                    "last_activity_at": item.get("last_activity_at") or (last_activity_dt.isoformat(sep=" ") if last_activity_dt else ""),
                })
            elif item.get("is_memo"):
                fact_memos.append(item)
                for c in item.get("memo_comments", []):
                    fact_memo_comments.append(c)

            fact_interactions.append({
                "entry_id": item.get("entry_id"),
                "project_id": item.get("project_id"),
                "meeting_id": item.get("meeting_id"),
                "company": item.get("company"),
                "event_type": "entry_closed" if item.get("is_closed") else "entry_open",
                "event_at": item.get("completed_at") or item.get("meeting_date") or "",
            })

        meetings_count = 0
        fact_meetings: List[Dict[str, Any]] = []
        meeting_rows_for_project = []
        for m in meetings_rows:
            pid = self._get_first_value(m, ["Project/ID", "Project"])
            if resolved_id and pid and pid != resolved_id:
                continue
            meeting_rows_for_project.append(m)

        for m in meeting_rows_for_project:
            mid = self._row_id(m)
            mdate = self._get_first_value(m, METRONOME_COLUMN_ALIASES["meeting_date"])
            meetings_count += 1
            fact_meetings.append({
                "meeting_id": mid,
                "meeting_date": mdate,
                "project_id": resolved_id,
                "project_title": resolved_title,
                "attending_company_ids": self._split_multi_values(self._get_first_value(m, METRONOME_COLUMN_ALIASES["meeting_attending_company_ids"])),
                "missing_company_ids": self._split_multi_values(self._get_first_value(m, METRONOME_COLUMN_ALIASES["meeting_missing_company_ids"])),
                "raw": m,
            })

        documents_for_project = [
            d for d in documents_rows
            if (not resolved_id or self._get_first_value(d, METRONOME_COLUMN_ALIASES["documents_project_id"]) in {"", resolved_id})
        ]

        project_start = parse_date(self._get_first_value(project_info, METRONOME_COLUMN_ALIASES["project_start"]))
        project_end = parse_date(self._get_first_value(project_info, METRONOME_COLUMN_ALIASES["project_end"]))
        if project_start and project_end and project_end > project_start:
            total_days = (project_end - project_start).days
            elapsed_days_raw = max((today - project_start).days, 0)
            elapsed_days = min(elapsed_days_raw, total_days)
            progress_percent = round((elapsed_days / total_days) * 100, 1)
            overrun_days = max(0, elapsed_days_raw - total_days)
            overrun_pct = round((overrun_days / total_days) * 100, 1) if total_days else 0.0
        else:
            total_days = 0
            elapsed_days_raw = 0
            elapsed_days = 0
            progress_percent = 0.0
            overrun_days = 0
            overrun_pct = 0.0

        total_tasks = len(fact_tasks)
        closed_tasks = sum(1 for t in fact_tasks if t.get("is_closed"))
        open_tasks = sum(1 for t in fact_tasks if t.get("is_open"))
        late_open_tasks = sum(1 for t in fact_tasks if t.get("is_late"))
        closure_rate = round((closed_tasks / total_tasks) * 100, 1) if total_tasks else 0.0
        on_time_closed = sum(1 for t in fact_tasks if t.get("is_closed_on_time"))
        on_time_closure_rate = round((on_time_closed / closed_tasks) * 100, 1) if closed_tasks else 0.0
        avg_task_age_open = round(sum(t.get("age_days", 0) for t in fact_tasks if t.get("is_open")) / max(1, open_tasks), 1) if open_tasks else 0.0
        avg_days_late = round(sum(t.get("days_late", 0) for t in fact_tasks if t.get("is_late")) / max(1, late_open_tasks), 1) if late_open_tasks else 0.0
        tasks_per_meeting = round(total_tasks / max(1, meetings_count), 2)
        memos_per_meeting = round(len(fact_memos) / max(1, meetings_count), 2)
        operational_progress_pct = round((closed_tasks / total_tasks) * 100, 1) if total_tasks else 0.0
        calendar_progress_pct = progress_percent
        progress_gap_pct = round(operational_progress_pct - calendar_progress_pct, 1)
        recent_threshold = today - timedelta(days=30)
        closure_velocity_30d = sum(1 for t in fact_tasks if t.get("is_closed") and t.get("done_date") and t.get("done_date") >= recent_threshold)
        project_health_score = max(0.0, min(100.0, round(0.45*closure_rate + 0.3*on_time_closure_rate + 0.25*max(0,100-(late_open_tasks*100/max(1,total_tasks))), 1)))

        kpi_project_summary = {
            "total_tasks": total_tasks,
            "open_tasks": open_tasks,
            "closed_tasks": closed_tasks,
            "late_open_tasks": late_open_tasks,
            "closure_rate": closure_rate,
            "on_time_closure_rate": on_time_closure_rate,
            "avg_task_age_open": avg_task_age_open,
            "avg_days_late": avg_days_late,
            "meetings_count": meetings_count,
            "tasks_per_meeting": tasks_per_meeting,
            "memos_per_meeting": memos_per_meeting,
            "calendar_progress_pct": calendar_progress_pct,
            "operational_progress_pct": operational_progress_pct,
            "progress_gap_pct": progress_gap_pct,
            "closure_velocity_30d": closure_velocity_30d,
            "project_health_score": project_health_score,
        }

        company_task_map: Dict[str, List[Dict[str, Any]]] = {}
        for t in fact_tasks:
            company_task_map.setdefault(clean_text(t.get("company")) or "Non renseigné", []).append(t)

        attendance_by_company: Dict[str, Dict[str, int]] = {}
        for m in fact_meetings:
            for cid in m.get("attending_company_ids", []):
                cname = clean_text(companies.get(cid, {}).get("Name")) or cid
                rec = attendance_by_company.setdefault(cname, {"attended": 0, "missing": 0})
                rec["attended"] += 1
            for cid in m.get("missing_company_ids", []):
                cname = clean_text(companies.get(cid, {}).get("Name")) or cid
                rec = attendance_by_company.setdefault(cname, {"attended": 0, "missing": 0})
                rec["missing"] += 1

        kpi_company_summary: List[Dict[str, Any]] = []
        kpi_reactivity_company: List[Dict[str, Any]] = []
        for cname, tasks in company_task_map.items():
            total = len(tasks)
            open_n = sum(1 for x in tasks if x.get("is_open"))
            late_open = sum(1 for x in tasks if x.get("is_late"))
            closed = sum(1 for x in tasks if x.get("is_closed"))
            on_time = sum(1 for x in tasks if x.get("is_closed_on_time"))
            delays = [x.get("days_late", 0) for x in tasks if x.get("is_late")]
            response = [x.get("response_delay_days") for x in tasks if x.get("response_delay_days") is not None]
            under3 = sum(1 for v in response if v <= 3)
            no_response = sum(1 for x in tasks if x.get("response_delay_days") is None)
            attend = attendance_by_company.get(cname, {"attended": 0, "missing": 0})
            meetings_total_for_company = attend["attended"] + attend["missing"]
            reminder_count = sum(1 for x in tasks if x.get("is_open") and x.get("age_days", 0) >= reminder_threshold_days)

            rec = {
                "company": cname,
                "total_tasks": total,
                "open_tasks": open_n,
                "late_open_tasks": late_open,
                "closure_rate": round((closed / total) * 100, 1) if total else 0.0,
                "on_time_closure_rate": round((on_time / closed) * 100, 1) if closed else 0.0,
                "avg_days_late": round(sum(delays)/len(delays),1) if delays else 0.0,
                "avg_first_response_delay_days": round(sum(response)/len(response),1) if response else None,
                "no_response_rate": round((no_response/total)*100,1) if total else 0.0,
                "meetings_attended": attend["attended"],
                "meetings_missing": attend["missing"],
                "attendance_rate": round((attend["attended"]/meetings_total_for_company)*100,1) if meetings_total_for_company else None,
                "reminder_count": reminder_count,
                "reminder_per_task_ratio": round(reminder_count/total,2) if total else 0.0,
            }
            reactivity_score = 100.0
            if rec["avg_first_response_delay_days"] is not None:
                reactivity_score -= min(40.0, rec["avg_first_response_delay_days"] * 4)
            reactivity_score -= min(35.0, rec["no_response_rate"] * 0.35)
            reactivity_score += min(20.0, (under3 / max(1, len(response))) * 20 if response else 0)
            rec["reactivity_score"] = round(max(0.0, min(100.0, reactivity_score)), 1)
            kpi_company_summary.append(rec)
            kpi_reactivity_company.append({
                "company": cname,
                "avg_first_response_delay_days": rec["avg_first_response_delay_days"],
                "reaction_under_3d_rate": round((under3/len(response))*100,1) if response else 0.0,
                "tasks_without_response_rate": rec["no_response_rate"],
                "avg_processing_days": round(sum((x.get("done_date")-x.get("request_date")).days for x in tasks if x.get("done_date") and x.get("request_date")) / max(1, sum(1 for x in tasks if x.get("done_date") and x.get("request_date"))),1) if any(x.get("done_date") and x.get("request_date") for x in tasks) else None,
                "reactivity_score": rec["reactivity_score"],
            })

        kpi_company_summary.sort(key=lambda x: (-x["open_tasks"], x["company"]))
        kpi_reactivity_company.sort(key=lambda x: (x["reactivity_score"] if x["reactivity_score"] is not None else 999, x["company"]))

        package_map: Dict[str, List[Dict[str, Any]]] = {}
        for t in fact_tasks:
            package_label = clean_text(t.get("package_label") or t.get("package_name") or "Non renseigné")
            package_map.setdefault(package_label, []).append(t)
        kpi_package_summary: List[Dict[str, Any]] = []
        for plabel, tasks in package_map.items():
            total = len(tasks)
            open_n = sum(1 for x in tasks if x.get("is_open"))
            late_open = sum(1 for x in tasks if x.get("is_late"))
            closed = sum(1 for x in tasks if x.get("is_closed"))
            memos_count = sum(1 for m in fact_memos if plabel in [clean_text(v) for v in m.get("memo_package_labels", [])])
            impacted_areas_count = len({a for x in tasks for a in x.get("area_names", []) if clean_text(a)})
            criticality = round(min(100.0, late_open*12 + open_n*3 + impacted_areas_count*5),1)
            kpi_package_summary.append({
                "package": plabel,
                "total_tasks": total,
                "open_tasks": open_n,
                "late_open_tasks": late_open,
                "closure_rate": round((closed/total)*100,1) if total else 0.0,
                "avg_days_late": round(sum(x.get("days_late",0) for x in tasks if x.get("is_late"))/max(1,late_open),1) if late_open else 0.0,
                "tasks_due_7d": sum(1 for x in tasks if x.get("days_to_deadline") is not None and 0 <= x.get("days_to_deadline") <= 7),
                "memos_count": memos_count,
                "impacted_areas_count": impacted_areas_count,
                "package_criticality_score": criticality,
            })
        kpi_package_summary.sort(key=lambda x: (-x["late_open_tasks"], -x["open_tasks"], x["package"]))

        zone_map: Dict[str, List[Dict[str, Any]]] = {}
        for t in fact_tasks:
            names = t.get("area_names") or ["Non renseigné"]
            for z in names:
                zone = clean_text(z) or "Non renseigné"
                zone_map.setdefault(zone, []).append(t)
        kpi_zone_summary: List[Dict[str, Any]] = []
        for zname, tasks in zone_map.items():
            total = len(tasks)
            open_n = sum(1 for x in tasks if x.get("is_open"))
            late_open = sum(1 for x in tasks if x.get("is_late"))
            closed = sum(1 for x in tasks if x.get("is_closed"))
            critical_packages_count = len({clean_text(x.get("package_label")) for x in tasks if x.get("is_late")})
            kpi_zone_summary.append({
                "zone": zname,
                "total_tasks": total,
                "open_tasks": open_n,
                "late_open_tasks": late_open,
                "closure_rate": round((closed/total)*100,1) if total else 0.0,
                "critical_packages_count": critical_packages_count,
            })
        kpi_zone_summary.sort(key=lambda x: (-x["late_open_tasks"], -x["open_tasks"], x["zone"]))

        kpi_project_progress = {
            "calendar_progress_pct": calendar_progress_pct,
            "operational_progress_pct": operational_progress_pct,
            "progress_gap_pct": progress_gap_pct,
            "start_date": project_start.isoformat() if project_start else "",
            "end_date": project_end.isoformat() if project_end else "",
            "elapsed_days": elapsed_days,
            "elapsed_days_raw": elapsed_days_raw,
            "total_days": total_days,
            "overrun_days": overrun_days,
            "overrun_pct": overrun_pct,
            "is_overrun": overrun_days > 0,
        }

        for task in fact_tasks:
            dline = task.get("deadline")
            deadline_dt = parse_date(dline) if dline else None
            is_done = bool(task.get("is_closed"))
            if is_done:
                timeline_status = "clos"
            elif deadline_dt and deadline_dt < reference_date:
                timeline_status = "rappel"
            else:
                timeline_status = "a_suivre"
            start_dt = task.get("request_date")
            if not start_dt and deadline_dt:
                start_dt = deadline_dt - timedelta(days=7)
            task["timeline_status"] = timeline_status
            task["timeline_start"] = start_dt.isoformat() if start_dt else ""
            task["timeline_end"] = deadline_dt.isoformat() if deadline_dt else ""
            task["company"] = self._normalize_company(task.get("company"))
            task["owner_full_name"] = self._normalize_owner(task.get("owner_full_name"))

        reminder_log: List[Dict[str, Any]] = []
        rid = 0
        for task in fact_tasks:
            if task.get("is_open") and task.get("age_days", 0) >= reminder_threshold_days:
                rid += 1
                reminder_log.append({
                    "reminder_id": f"theoretical-{rid}",
                    "project_id": task.get("project_id"),
                    "meeting_id": task.get("meeting_id"),
                    "entry_id": task.get("entry_id"),
                    "company_id": task.get("package_company_id") or "",
                    "company_name": task.get("company") or "Non renseigné",
                    "sent_at": "",
                    "sent_by_user_id": "",
                    "sent_by_user_name": "",
                    "channel": "",
                    "reminder_type": "theoretical",
                    "template_name": "",
                    "status": "pending",
                })
            txt = clean_text(task.get("last_comment_text")).lower()
            if txt and ("relance" in txt or "rappel" in txt):
                rid += 1
                reminder_log.append({
                    "reminder_id": f"real-{rid}",
                    "project_id": task.get("project_id"),
                    "meeting_id": task.get("meeting_id"),
                    "entry_id": task.get("entry_id"),
                    "company_id": task.get("package_company_id") or "",
                    "company_name": task.get("company") or "Non renseigné",
                    "sent_at": task.get("last_comment_date") or task.get("last_activity_at") or "",
                    "sent_by_user_id": task.get("last_comment_user_id") or "",
                    "sent_by_user_name": task.get("last_comment_user_name") or "",
                    "channel": "memo_comment",
                    "reminder_type": "real",
                    "template_name": "",
                    "status": "sent",
                })

        reminder_count_theoretical = sum(1 for r in reminder_log if r.get("reminder_type") == "theoretical")
        reminder_count_real = sum(1 for r in reminder_log if r.get("reminder_type") == "real")
        for c in kpi_company_summary:
            cname = c.get("company")
            c["reminder_count_theoretical"] = sum(1 for r in reminder_log if r.get("reminder_type") == "theoretical" and r.get("company_name") == cname)
            c["reminder_count_real"] = sum(1 for r in reminder_log if r.get("reminder_type") == "real" and r.get("company_name") == cname)

        open_tasks_today = [t for t in fact_tasks if t.get("is_open")]
        overdue_tasks_today = [t for t in open_tasks_today if t.get("deadline") and parse_date(t.get("deadline")) and parse_date(t.get("deadline")) < today]
        reminder_tasks_today: List[Dict[str, Any]] = []
        reminder_tasks_cumulative: List[Dict[str, Any]] = []
        followup_tasks_today: List[Dict[str, Any]] = []
        open_not_due_tasks_today: List[Dict[str, Any]] = []
        for task in fact_tasks:
            deadline_txt = task.get("deadline")
            deadline_dt = parse_date(deadline_txt) if deadline_txt else None
            if not deadline_dt:
                continue
            pivot_dt = task.get("done_date") if task.get("is_closed") and task.get("done_date") else today
            delta_biz_cum = self._business_day_delta(deadline_dt, pivot_dt)
            if delta_biz_cum < 0 and abs(delta_biz_cum) >= 7:
                reminder_tasks_cumulative.append(task)

        for task in open_tasks_today:
            deadline_txt = task.get("deadline")
            deadline_dt = parse_date(deadline_txt) if deadline_txt else None
            if not deadline_dt:
                open_not_due_tasks_today.append(task)
                continue
            delta_biz = self._business_day_delta(deadline_dt, today)
            if delta_biz < 0:
                if abs(delta_biz) >= 7:
                    reminder_tasks_today.append(task)
                else:
                    followup_tasks_today.append(task)
            else:
                open_not_due_tasks_today.append(task)
                if delta_biz <= 7:
                    followup_tasks_today.append(task)

        def group_company(rows_in: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
            counts: Dict[str, int] = {}
            logos: Dict[str, str] = {}
            for task in rows_in:
                cname = self._normalize_company(task.get("company"))
                counts[cname] = counts.get(cname, 0) + 1
                logos[cname] = company_logos_by_name.get(cname, "")
            return [
                {"name": n, "count": c, "logo": logos.get(n, "")}
                for n, c in sorted(counts.items(), key=lambda x: (-x[1], x[0]))
            ]

        reactivity_by_company: List[Dict[str, Any]] = []
        delays: Dict[str, List[int]] = {}
        for task in fact_tasks:
            if not task.get("is_closed"):
                continue
            deadline_txt = task.get("deadline")
            deadline_dt = parse_date(deadline_txt) if deadline_txt else None
            done_dt = task.get("done_date")
            if not deadline_dt or not done_dt:
                continue
            cname = self._normalize_company(task.get("company"))
            delays.setdefault(cname, []).append((done_dt - deadline_dt).days)
        for cname, vals in sorted(delays.items(), key=lambda x: x[0]):
            reactivity_by_company.append({
                "name": cname,
                "avg_gap_days": round(sum(vals) / max(1, len(vals)), 1),
                "closed_count": len(vals),
                "logo": company_logos_by_name.get(cname, ""),
            })
        open_entries_count = rows_filtered_by_title + rows_filtered_by_id
        if project_info and match_debug.get("match_score", 0) >= 80:
            confidence_level = "high"
        elif open_entries_count > 0 or match_debug.get("resolution_mode") == "entries_fallback":
            confidence_level = "medium"
        else:
            confidence_level = "low"

        warning = bool(match_debug.get("resolution_mode") == "entries_fallback" or not project_info)
        warning_message = "Projet trouvé via entrées METRONOME (fallback)" if warning else ""

        project_display_name = resolved_title or target
        project_id = resolved_id or self._row_id(project_info)
        project_image = self._get_first_value(project_info, METRONOME_COLUMN_ALIASES["project_image"])
        project_description = self._get_first_value(project_info, METRONOME_COLUMN_ALIASES["project_description"])
        return {
            "ok": True,
            "project_name": project_display_name,
            "project_id": project_id,
            "project_image": project_image,
            "project_description": project_description,
            "confidence_level": confidence_level,
            "warning": warning,
            "warning_message": warning_message,
            "match_debug": match_debug,
            "kpis_meeting_simple": meeting_simple_kpis,
            "kpis": {
                "open_topics": len(open_tasks_today),
                "overdue_topics": len(overdue_tasks_today),
                "by_company": by_company,
                "by_package": count_by("lot"),
                "by_meeting": by_meeting,
            },
            "kpis_pilotage": {
                "reminders_open_count": len(reminder_tasks_today),
                "followups_open_count": len(open_not_due_tasks_today),
                "reference_date": reference_date.isoformat(),
                "reference_date_text": reference_date_text,
                "reminders_by_company": group_company(reminder_tasks_today),
                "open_tasks_by_company": open_tasks_by_company,
                "average_processing_days_by_company": average_processing_days_by_company,
                "rappels_ouverts_a_date": len(reminder_tasks_today),
                "a_suivre_ouverts": len(open_not_due_tasks_today),
                "date_reference": reference_date.isoformat(),
                "rappels_cumules_par_entreprise": group_company(reminder_tasks_today),
                "project_company_views": {
                    "open": group_company(open_not_due_tasks_today),
                    "followup": group_company(followup_tasks_today),
                    "reminder": group_company(reminder_tasks_cumulative),
                },
                "reminders_open_items": [
                    {
                        "company": self._normalize_company(t.get("company")),
                        "perimetre": self._normalize_zone((t.get("area_names") or [""])[0] if (t.get("area_names") or []) else ""),
                        "lot": self._normalize_lot(t.get("package_name") or t.get("package_label")),
                        "task": clean_text(t.get("entry_title")),
                        "deadline": clean_text(t.get("deadline")),
                        "owner": self._normalize_owner(t.get("owner_full_name")),
                    }
                    for t in sorted(reminder_tasks_today, key=lambda x: (self._normalize_zone((x.get("area_names") or [""])[0] if (x.get("area_names") or []) else ""), self._normalize_lot(x.get("package_name") or x.get("package_label"),), clean_text(x.get("entry_title"))))
                ],
                "reactivity_by_company": reactivity_by_company,
            },
            "analytics": {
                "fact_tasks": fact_tasks,
                "fact_memos": fact_memos,
                "fact_memo_comments": fact_memo_comments,
                "fact_meetings": fact_meetings,
                "fact_interactions": fact_interactions,
                "kpi_project_summary": kpi_project_summary,
                "kpi_company_summary": kpi_company_summary,
                "kpi_package_summary": kpi_package_summary,
                "kpi_zone_summary": kpi_zone_summary,
                "kpi_project_progress": kpi_project_progress,
                "kpi_reactivity_company": kpi_reactivity_company,
                "reminder_log": reminder_log,
                "reminder_count_theoretical": reminder_count_theoretical,
                "reminder_count_real": reminder_count_real,
                "documents_count": len(documents_for_project),
            },
            "rows": rows,
            "missing_files": cache.get("missing", []),
            "loaded_at": cache.get("loaded_at"),
        }


class PointageService:
    def __init__(self, store_file: Path) -> None:
        self.store_file = store_file
        self._lock = Lock()

    def _load(self) -> Dict[str, Any]:
        if not self.store_file.exists():
            return {"projects": {}}
        try:
            return json.loads(self.store_file.read_text(encoding="utf-8"))
        except Exception:
            return {"projects": {}}

    def _save(self, data: Dict[str, Any]) -> None:
        self.store_file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    @staticmethod
    def _project_key(affaire_id: str, affaire_name: str) -> str:
        aid = clean_text(affaire_id)
        if aid:
            return f"id::{aid}"
        return f"name::{slugify(affaire_name)}"

    @staticmethod
    def _decode_csv(raw: bytes) -> str:
        for enc in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                return raw.decode(enc)
            except Exception:
                continue
        return raw.decode("utf-8", errors="ignore")

    @staticmethod
    def _csv_rows(text: str) -> List[Dict[str, str]]:
        first = text.splitlines()[0] if text.splitlines() else ""
        delimiter = ";" if first.count(";") >= first.count(",") and first.count(";") > 0 else ("	" if "	" in first else ",")
        return list(csv.DictReader(io.StringIO(text), delimiter=delimiter))

    @staticmethod
    def _find_col(row: Dict[str, Any], names: List[str]) -> str:
        for n in names:
            for k in row.keys():
                if slugify(k) == slugify(n) or slugify(n) in slugify(k):
                    return k
        return ""

    @staticmethod
    def _parse_outline_level(raw_value: Any) -> Optional[int]:
        text = clean_text(raw_value)
        if not text:
            return None
        m = re.search(r"\d+", text)
        if not m:
            return None
        try:
            return max(0, int(m.group(0)))
        except Exception:
            return None

    @staticmethod
    def _parse_duration_to_hours(raw_value: Any, default_hours_per_day: float = 8.0) -> Tuple[float, float, str]:
        text = clean_text(raw_value)
        if not text:
            return 0.0, 0.0, ""
        compact = text.lower().replace(",", ".")
        m = re.search(r"(-?\d+(?:\.\d+)?)", compact)
        if not m:
            return 0.0, 0.0, ""
        value = clean_number(m.group(1))
        if re.search(r"\b(j|jr|jrs|jour|jours|day|days)\b", compact):
            return value * default_hours_per_day, value, "days"
        if re.search(r"\b(min|mn|minute|minutes)\b", compact):
            return value / 60.0, value, "minutes"
        if re.search(r"\b(h|hr|hrs|heure|heures|hour|hours)\b", compact):
            return value, value, "hours"
        return value, value, "hours"

    def parse_planning(self, raw: bytes) -> List[Dict[str, Any]]:
        text = self._decode_csv(raw)
        rows = self._csv_rows(text)
        if not rows:
            return []
        first = rows[0]
        c_id = self._find_col(first, ["id", "task id", "id tache", "uid"])
        c_name = self._find_col(first, ["task name", "nom", "name", "tache"])
        c_start = self._find_col(first, ["start", "start date", "start1", "début (f)", "debut (f)", "debut", "début", "date debut", "date début"])
        c_end = self._find_col(first, ["finish", "finish date", "finish1", "fin (g)", "fin", "date fin", "end", "end date"])
        c_work = self._find_col(first, ["work", "charge", "planned hours", "travail", "duree", "durée", "duration"])
        c_duration = self._find_col(first, ["duration", "duree", "durée", "planned duration"])
        c_owner = self._find_col(first, ["owner", "resource", "nom de ressources", "nom de ressource", "responsable", "ressource"])
        c_pct = self._find_col(first, ["%", "% acheve", "% achevé", "% complete", "percent", "percentage complete", "percent complete", "acheve", "achevé", "avancement", "pourcentage_acheve", "pourcentage acheve"])
        c_outline = self._find_col(first, ["outline level", "outline", "niveau", "level", "wbs", "indent"])
        c_summary = self._find_col(first, ["summary", "is summary", "recap", "récap"])
        c_pred = self._find_col(first, ["predecessors", "pred", "predecesseurs", "prédécesseurs"])
        c_unit_cost = self._find_col(first, ["variation_de_cout", "variation de cout", "variation_de_coût", "variation de coût", "variation_de_coût (p)", "variation de coût (p)", "unit cost", "cout unitaire", "coût unitaire", "rate"])

        tasks: List[Dict[str, Any]] = []
        outlines: List[int] = []
        current_level = "Général"
        current_process = ""
        parent_stack: List[str] = []
        for idx, r in enumerate(rows):
            name = clean_text(r.get(c_name)) if c_name else ""
            if not name:
                continue
            name_slug = slugify(name)
            if "synthese-technique-niveau" in name_slug or re.search(r"niveau\s+[a-z0-9-]+", name_slug):
                current_level = name
                current_process = ""
            if name_slug.startswith("process"):
                current_process = name
            task_id = clean_text(r.get(c_id)) if c_id else ""
            if not task_id:
                task_id = f"task-{idx+1}-{slugify(name)[:24]}"
            start = MetronomeService._parse_date_only(clean_text(r.get(c_start))) if c_start else None
            end = MetronomeService._parse_date_only(clean_text(r.get(c_end))) if c_end else None
            duration_raw = clean_text(r.get(c_duration)) if c_duration else ""
            work_raw = clean_text(r.get(c_work)) if c_work else ""
            duration_source = duration_raw or work_raw
            planned_hours, planned_duration_value, planned_duration_unit = self._parse_duration_to_hours(duration_source)
            owner = clean_text(r.get(c_owner)) if c_owner else ""
            pct = clean_number(r.get(c_pct)) if c_pct else 0.0
            outline_level = self._parse_outline_level(r.get(c_outline)) if c_outline else None
            if outline_level is None:
                outline_level = 0 if name_slug.startswith("valhubert-") else (1 if name_slug.startswith("process") else 2)
            outlines.append(outline_level)

            explicit_summary = clean_text(r.get(c_summary)).lower() if c_summary else ""
            is_summary_explicit = explicit_summary in {"1", "true", "yes", "oui", "summary", "recap", "récap"}
            is_structure = is_summary_explicit or name_slug.startswith("process") or "synthese-technique" in name_slug
            is_cst = "cst" in slugify(owner) or "cst" in name_slug
            is_cet = slugify(owner) in {"cet"} or slugify(owner).startswith("cet-")
            stable_key = f"{task_id}|{slugify(name)}|{slugify(current_level)}|{idx}"

            while len(parent_stack) > outline_level:
                parent_stack.pop()
            parent_task_id = parent_stack[-1] if parent_stack else ""
            parent_stack.append(task_id)

            tasks.append({
                "task_id": task_id,
                "stable_key": stable_key,
                "name": name,
                "level": current_level,
                "process_label": current_process,
                "level_label": current_level,
                "parent_task_id": parent_task_id,
                "outline_level": outline_level,
                "depth": outline_level,
                "owner": owner or "Non attribué",
                "start": start.isoformat() if start else "",
                "end": end.isoformat() if end else "",
                "planned_hours": planned_hours,
                "planned_duration_value": planned_duration_value,
                "planned_duration_unit": planned_duration_unit,
                "unit_cost": clean_number(r.get(c_unit_cost)) if c_unit_cost else 0.0,
                "cost_variation": clean_number(r.get(c_unit_cost)) if c_unit_cost else 0.0,
                "predecessors": clean_text(r.get(c_pred)) if c_pred else "",
                "csv_progress": max(0.0, min(100.0, pct)),
                "is_summary": is_structure,
                "is_actionable": not is_structure,
                "is_cst": is_cst,
                "is_cet": is_cet,
            })

        for i, task in enumerate(tasks):
            next_outline = outlines[i + 1] if i + 1 < len(outlines) else -1
            if next_outline > int(task.get("outline_level", 0)):
                task["is_summary"] = True
                task["is_actionable"] = False

        return tasks

    def _ensure_project(self, key: str) -> Dict[str, Any]:
        data = self._load()
        proj = data.setdefault("projects", {}).setdefault(key, {
            "planning_tasks": [],
            "pointage": {"__cetMembers": "", "__cstRate": 80},
            "workState": {"expanded_levels": []},
            "updated_at": now_iso(),
        })
        return data

    def import_planning(self, project_key: str, raw: bytes) -> Dict[str, Any]:
        tasks = self.parse_planning(raw)
        with self._lock:
            data = self._ensure_project(project_key)
            proj = data["projects"][project_key]
            existing_pointage = proj.get("pointage", {})
            remapped: Dict[str, Any] = {k: v for k, v in existing_pointage.items() if k.startswith("__")}
            by_task = {t.get("task_id"): t for t in proj.get("planning_tasks", [])}
            by_stable = {t.get("stable_key"): t for t in proj.get("planning_tasks", [])}
            for nt in tasks:
                old = by_task.get(nt.get("task_id")) or by_stable.get(nt.get("stable_key"))
                if old and old.get("task_id") in existing_pointage:
                    remapped[nt["task_id"]] = existing_pointage[old["task_id"]]
            proj["planning_tasks"] = tasks
            proj["pointage"] = remapped
            proj["updated_at"] = now_iso()
            self._save(data)
        return self.get_project_data(project_key)

    def save_pointage(self, project_key: str, pointage_patch: Dict[str, Any], work_state: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        with self._lock:
            data = self._ensure_project(project_key)
            proj = data["projects"][project_key]
            p = proj.setdefault("pointage", {"__cetMembers": "", "__cstRate": 80})
            for k, v in pointage_patch.items():
                p[k] = v
            if work_state is not None:
                proj["workState"] = work_state
            proj["updated_at"] = now_iso()
            self._save(data)
        return self.get_project_data(project_key)

    def import_suivi(self, project_key: str, raw_json: bytes) -> Dict[str, Any]:
        payload = json.loads(raw_json.decode("utf-8"))
        pointage = payload.get("pointage", {})
        work_state = payload.get("workState", {})
        return self.save_pointage(project_key, pointage, work_state=work_state)

    def export_suivi(self, project_key: str) -> Dict[str, Any]:
        data = self.get_project_data(project_key)
        return {"exportedAt": now_iso(), "pointage": data.get("pointage", {}), "workState": data.get("workState", {})}

    def compute_tasks(self, planning_tasks: List[Dict[str, Any]], pointage: Dict[str, Any]) -> List[Dict[str, Any]]:
        cet_members = [clean_text(x) for x in clean_text(pointage.get("__cetMembers", "")).split(",") if clean_text(x)]
        cst_rate = clean_number(pointage.get("__cstRate", 80))
        today = datetime.now().date()
        out: List[Dict[str, Any]] = []
        for t in planning_tasks:
            rec = pointage.get(t["task_id"], {}) if isinstance(pointage.get(t["task_id"]), dict) else {}
            progress = clean_number(rec.get("progress", t.get("csv_progress", 0)))
            actual_end = MetronomeService._parse_date_only(clean_text(rec.get("actualEnd")))
            cet_map = rec.get("cet", {}) if isinstance(rec.get("cet"), dict) else {}
            cet_progress_vals = []
            cet_dates: List[date] = []
            if t.get("is_cet") and cet_members:
                for m in cet_members:
                    cm = cet_map.get(m, {}) if isinstance(cet_map.get(m), dict) else {}
                    cp = clean_number(cm.get("progress", 0))
                    ca = MetronomeService._parse_date_only(clean_text(cm.get("actualEnd")))
                    if ca:
                        cp = 100.0
                        cet_dates.append(ca)
                    cet_progress_vals.append(max(0.0, min(100.0, cp)))
                if not rec.get("progress") and cet_progress_vals:
                    progress = sum(cet_progress_vals) / len(cet_progress_vals)
                if not actual_end and cet_dates:
                    actual_end = max(cet_dates)
            if actual_end:
                progress = 100.0
            progress = max(0.0, min(100.0, progress))
            start = MetronomeService._parse_date_only(t.get("start", ""))
            end = MetronomeService._parse_date_only(t.get("end", ""))
            if actual_end and end:
                delay_days = max(0, (actual_end - end).days)
            elif (not actual_end) and end and progress < 100 and today > end:
                delay_days = (today - end).days
            else:
                delay_days = 0
            if progress >= 100 or actual_end:
                status = "past"
            elif start and today < start:
                status = "future"
            else:
                status = "current"
            planned_hours = clean_number(t.get("planned_hours"))
            cost_variation = clean_number(t.get("cost_variation"))
            planned_cost = cost_variation if cost_variation != 0 else (planned_hours * cst_rate if t.get("is_cst") else 0.0)
            actual_cost = planned_cost * (progress / 100.0)
            out.append({
                **t,
                "progress": round(progress, 1),
                "actualEnd": actual_end.isoformat() if actual_end else clean_text(rec.get("actualEnd")),
                "cet": cet_map,
                "cetMembers": cet_members,
                "status": status,
                "delayDays": int(delay_days),
                "plannedCostCst": round(planned_cost, 2),
                "actualCostCst": round(actual_cost, 2),
                "plannedHours": round(planned_hours, 2),
                "cstRate": cst_rate,
                "isLate": bool(delay_days > 0),
            })
        return out

    def get_project_data(self, project_key: str) -> Dict[str, Any]:
        data = self._load()
        proj = data.get("projects", {}).get(project_key, {"planning_tasks": [], "pointage": {"__cetMembers": "", "__cstRate": 80}, "workState": {"expanded_levels": []}})
        tasks = self.compute_tasks(proj.get("planning_tasks", []), proj.get("pointage", {}))
        return {
            "project_key": project_key,
            "tasks": tasks,
            "pointage": proj.get("pointage", {}),
            "workState": proj.get("workState", {}),
            "updated_at": proj.get("updated_at", ""),
        }


service = FinanceService(WORKBOOK_PATH, SHEET_NAME, CACHE_FILE)
metronome_service = MetronomeService(METRONOME_BASE_PATH)
pointage_service = PointageService(POINTAGE_STORE_FILE)
boond_service = BoondService(BOOND_CACHE_DB_PATH)


def pointage_finance_summary(affaire_id: str, commande_ht: float) -> Dict[str, float]:
    try:
        data = pointage_service.get_project_data(PointageService._project_key(affaire_id, ""))
        tasks = data.get("tasks", []) if isinstance(data, dict) else []
    except Exception:
        tasks = []
    actionable = [t for t in tasks if not bool(t.get("is_summary"))]
    planned = sum(clean_number(t.get("plannedCostCst")) for t in actionable)
    actual = sum(clean_number(t.get("actualCostCst")) for t in actionable)
    progress_ratio = (actual / planned) if planned > 0 else 0.0
    progress_amount = clean_number(commande_ht) * progress_ratio if clean_number(commande_ht) > 0 else 0.0
    return {
        "pointage_cost_planned_cst": round(planned, 2),
        "pointage_cost_actual_cst": round(actual, 2),
        "pointage_progress_ratio": round(progress_ratio, 4),
        "pointage_progress_amount": round(progress_amount, 2),
    }




def boond_imputations_html() -> str:
    return """<!doctype html>
<html lang='fr'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>
<title>Imputation des temps</title>
<style>
body{font-family:Inter,Arial,sans-serif;background:#f5f7fb;margin:0;color:#122033}.wrap{max-width:980px;margin:24px auto;padding:0 16px}
.card{background:#fff;border:1px solid #dfe5ef;border-radius:14px;padding:16px;margin-bottom:12px}.kpis{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}
.small{color:#6e7a90;font-size:13px}.value{font-size:30px;font-weight:800}.warn{padding:10px;border-radius:10px;background:#fff7ef;border:1px solid #f1d4a4}
table{width:100%;border-collapse:collapse}th,td{padding:8px;border-bottom:1px solid #eef2f7;text-align:left}
.bar{height:10px;background:#eef2f7;border-radius:999px;overflow:hidden}.fill{height:100%;background:#ef8d00}
</style></head><body><div class='wrap'>
<div class='card'><h2>Imputation des temps</h2><div id='state' class='small'>Chargement…</div></div>
<div id='content' style='display:none'>
<div class='card'><div class='small'>Projet METRONOME</div><div id='inputName'></div><div class='small' style='margin-top:6px'>Projet BOOND matché</div><div id='matched'></div></div>
<div class='card kpis'><div><div class='small'>Jours imputés cumulés</div><div id='kpiDays' class='value'>-</div></div><div><div class='small'>Coût cumulé</div><div id='kpiCost' class='value'>-</div></div></div>
<div id='warning' class='card warn' style='display:none'></div>
<div class='card'><h3>Répartition mensuelle</h3><div id='chart'></div><div id='table'></div></div>
</div></div>
<script>
function esc(v){return String(v??'').replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c]||c));}
function fmt(v){const n=Number(v||0);return n.toLocaleString('fr-FR',{maximumFractionDigits:2});}
(async function(){
  const params=new URLSearchParams(location.search); const project=params.get('project_name')||'';
  const state=document.getElementById('state'); const content=document.getElementById('content');
  if(!project){state.textContent='Paramètre project_name manquant.';return;}
  try{
    const r=await fetch('/api/boond/imputations/by-project?project_name='+encodeURIComponent(project));
    const d=await r.json(); if(!r.ok) throw new Error(d.detail||'Erreur API');
    state.style.display='none'; content.style.display='block';
    document.getElementById('inputName').textContent=d.input_project_name||'-';
    if(!d.matched_project){document.getElementById('matched').textContent='Aucun match BOOND';document.getElementById('kpiDays').textContent='0';document.getElementById('kpiCost').textContent='-';return;}
    document.getElementById('matched').textContent=`${d.matched_project.project_reference} (score ${d.matched_project.match_score})`;
    document.getElementById('kpiDays').textContent=fmt(d.totals.total_days);
    document.getElementById('kpiCost').textContent=d.totals.total_cost==null?'Indisponible':fmt(d.totals.total_cost)+' €';
    if(d.totals.cost_status!=='ok'){const w=document.getElementById('warning');w.style.display='block';w.textContent=d.message||'Coût indisponible faute de taux journalier.';}
    const rows=d.by_month||[]; const max=Math.max(1,...rows.map(x=>Number(x.days||0)));
    document.getElementById('chart').innerHTML=rows.map(x=>`<div style='margin:8px 0'><div class='small'>${esc(x.month)} — ${fmt(x.days)} j</div><div class='bar'><div class='fill' style='width:${(Number(x.days||0)/max)*100}%'></div></div></div>`).join('')||"<div class='small'>Aucune donnée mensuelle</div>";
    document.getElementById('table').innerHTML=rows.length?`<table><thead><tr><th>Mois</th><th>Jours</th><th>Coût</th></tr></thead><tbody>${rows.map(x=>`<tr><td>${esc(x.month)}</td><td>${fmt(x.days)}</td><td>${x.cost==null?'-':fmt(x.cost)+' €'}</td></tr>`).join('')}</tbody></table>`:"";
  }catch(e){state.textContent=e.message||'Erreur';}
})();
</script></body></html>"""


@app.get("/api/boond/imputations/by-project", response_class=JSONResponse)
def api_boond_imputation_by_project(project_name: str = Query(...), refresh: bool = Query(default=False)):
    pname = clean_text(project_name)
    if not pname:
        raise HTTPException(status_code=400, detail="project_name obligatoire")
    try:
        return boond_service.get_project_imputation_summary(project_name=pname, refresh=bool(refresh))
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/boond-imputations", response_class=HTMLResponse)
def boond_imputations_page():
    return HTMLResponse(boond_imputations_html())

def landing_html() -> str:
    return """<!doctype html>
<html lang='fr'>
<head>
<meta charset='utf-8'>
<meta name='viewport' content='width=device-width, initial-scale=1'>
<title>Gestion Affaire - Accueil</title>
<style>
:root{--panel:#fff;--line:#dfe5ef;--ink:#122033;--muted:#6e7a90;--accent:#ef8d00;--shadow:0 16px 48px rgba(18,32,51,.09)}
*{box-sizing:border-box}body{margin:0;font-family:Inter,Segoe UI,Arial,sans-serif;background:linear-gradient(180deg,#f5f7fb 0%,#eef2f7 100%);color:var(--ink)}
.wrap{max-width:1260px;margin:30px auto 52px;padding:0 20px}.hero,.card{background:var(--panel);border:1px solid rgba(20,32,51,.05);border-radius:24px;box-shadow:var(--shadow)}
.hero{padding:28px}.hero-top{display:flex;align-items:center;gap:16px}.logo{width:56px;height:56px;object-fit:contain;border-radius:10px;background:#fff7ef;border:1px solid #f8ddb4;padding:6px}
.eyebrow{font-size:12px;font-weight:800;color:var(--accent);text-transform:uppercase;letter-spacing:.14em}h1{margin:8px 0 12px;font-size:44px;line-height:1.05}.sub{font-size:20px;color:var(--muted);max-width:930px}
.selector{margin-top:22px;display:grid;grid-template-columns:2fr 1fr;gap:12px}.search{height:52px;border:1px solid var(--line);border-radius:14px;padding:0 14px;font-size:16px;width:100%}
.badge{display:inline-flex;align-items:center;justify-content:center;padding:0 14px;height:52px;border-radius:14px;background:#f9fbff;color:#41547b;font-weight:700;border:1px solid var(--line)}
.grid{margin-top:18px;display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px}.card{padding:20px;display:flex;flex-direction:column;justify-content:space-between;min-height:220px}.card h3{margin:0;font-size:21px}.card p{margin:10px 0 18px;color:var(--muted);font-size:14px;min-height:44px}.card .btn{width:100%}
.btn{display:inline-flex;align-items:center;justify-content:center;gap:8px;height:44px;padding:0 16px;border-radius:12px;text-decoration:none;border:none;font-weight:800;cursor:pointer}
.btn.primary,.btn.dark{background:var(--accent);color:#fff}.btn.disabled{background:#ecf0f7;color:#8a94a8;cursor:not-allowed}
.state{margin-top:14px;padding:14px 16px;border-radius:14px;background:#f8faff;border:1px solid var(--line);color:#4f5d78;font-weight:600}
@media (max-width:980px){.grid{grid-template-columns:repeat(2,1fr)}.selector{grid-template-columns:1fr}}@media (max-width:640px){h1{font-size:34px}.grid{grid-template-columns:1fr}}
</style>
</head>
<body>
  <div class='wrap'>
    <section class='hero'>
      <div class='hero-top'>
        <img class='logo' src='/assets/logo-tempo' alt='Logo Tempo'>
        <div><div class='eyebrow'>Gestion affaire</div><h1>Accueil multiprojets</h1></div>
      </div>
      <div class='sub'>Choisissez une affaire en saisissant ses premières lettres, puis accédez directement aux modules disponibles.</div>
      <div class='selector'>
        <input id='projectSearch' class='search' list='projectList' placeholder='Tapez les premières lettres de l'affaire'>
        <datalist id='projectList'></datalist>
        <div id='projectBadge' class='badge'>Aucune affaire sélectionnée</div>
      </div>
      <div id='state' class='state'>Sélectionnez une affaire pour activer les modules.</div>
    </section>

    <section class='grid'>
      <article class='card'>
        <h3>Tableau de bord</h3><p>Porte d'entrée de pilotage de l'affaire.</p>
        <a id='dashboardLink' class='btn disabled' href='javascript:void(0)' aria-disabled='true'>Ouvrir le tableau de bord</a>
      </article>
      <article class='card'>
        <h3>Finances</h3><p>Cockpit financier détaillé de l'affaire.</p>
        <a id='financeLink' class='btn disabled' href='javascript:void(0)' aria-disabled='true'>Ouvrir Finances</a>
      </article>
      <article class='card'><h3>Gestion de projet</h3><p>Planification, jalons et coordination.</p><a id='pmLink' class='btn disabled' href='javascript:void(0)' aria-disabled='true'>Ouvrir</a></article>
      <article class='card'><h3>Imputation</h3><p>Suivi des temps et affectations.</p><button class='btn disabled' disabled>Ouvrir</button></article>
    </section>
  </div>
<script>
const state={projects:[],selectedId:'',selectedLabel:''};
function esc(v){return String(v||'').replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]));}
function setModuleLink(id,base){const el=document.getElementById(id);if(state.selectedId){el.className='btn primary';el.href=`/${base}?affaire_id=${encodeURIComponent(state.selectedId)}`;el.removeAttribute('aria-disabled');}else{el.className='btn disabled';el.href='javascript:void(0)';el.setAttribute('aria-disabled','true');}}
function updateUi(){document.getElementById('projectBadge').textContent=state.selectedLabel?`Affaire : ${state.selectedLabel}`:'Aucune affaire sélectionnée';document.getElementById('state').textContent=state.selectedLabel?`Vous naviguez sur l'affaire ${state.selectedLabel}.`:'Sélectionnez une affaire pour activer les modules.';setModuleLink('financeLink','finance');setModuleLink('dashboardLink','dashboard');setModuleLink('pmLink','gestion-projet');if(state.selectedId){localStorage.setItem('selectedAffaireId',state.selectedId);} }
async function loadProjects(){try{const res=await fetch('/api/finance/affaires');const data=await res.json();state.projects=(data.items||[]).map(x=>({id:x.affaire_id,label:x.display_name})).filter(x=>x.id&&x.label).sort((a,b)=>a.label.localeCompare(b.label,'fr'));}catch(_){state.projects=[];}
const list=document.getElementById('projectList');list.innerHTML=state.projects.map(p=>`<option value="${esc(p.label)}"></option>`).join('');const savedId=localStorage.getItem('selectedAffaireId')||'';const selected=state.projects.find(x=>x.id===savedId);if(selected){state.selectedId=selected.id;state.selectedLabel=selected.label;document.getElementById('projectSearch').value=selected.label;}updateUi();}
document.getElementById('projectSearch').addEventListener('input',ev=>{const q=(ev.target.value||'').trim().toLowerCase();const selected=state.projects.find(p=>p.label.toLowerCase()===q)||state.projects.find(p=>p.label.toLowerCase().startsWith(q));state.selectedId=selected?.id||'';state.selectedLabel=selected?.label||'';updateUi();});
loadProjects();
</script>
</body>
</html>"""



def finance_html() -> str:
    return """
<!doctype html>
<html lang='fr'>
<head>
<meta charset='utf-8'>
<meta name='viewport' content='width=device-width, initial-scale=1'>
<title>Finance Cockpit</title>
<style>
:root{--bg:#eef2f7;--panel:#fff;--panel2:#f7f9fc;--line:#dfe5ef;--ink:#122033;--muted:#6e7a90;--blue:#ef8d00;--green:#1d9a5b;--amber:#c48716;--red:#c84c4c;--shadow:0 14px 40px rgba(19,31,53,.08)}
*{box-sizing:border-box}body{margin:0;font-family:Inter,Segoe UI,Arial,sans-serif;background:linear-gradient(180deg,#f5f7fb 0%,#eef2f7 100%);color:var(--ink)}
.container{max-width:1440px;margin:22px auto 36px;padding:0 18px}.topbar,.hero,.section,.kpis{background:var(--panel);border:1px solid rgba(22,34,51,.04);box-shadow:var(--shadow);border-radius:24px}
.topbar{display:flex;gap:16px;align-items:center;padding:18px 20px;position:sticky;top:14px;z-index:10}.brand{min-width:200px}.brand h1{margin:0;font-size:36px;line-height:1}.brand p{margin:6px 0 0;color:var(--muted);font-size:14px}
.controls{display:flex;gap:12px;align-items:center;flex:1;flex-wrap:wrap}.search,.select{height:48px;border-radius:14px;border:1px solid var(--line);background:#fff;padding:0 14px;color:var(--ink);font-size:15px}.search{min-width:280px;flex:1}.select{min-width:340px;flex:1}
.btn{height:48px;border:none;border-radius:14px;padding:0 16px;font-weight:700;cursor:pointer}.btn.primary{background:var(--blue);color:#fff}.btn.dark{background:#ef8d00;color:#fff}
.badge{display:inline-flex;align-items:center;gap:8px;padding:10px 12px;border-radius:999px;background:var(--panel2);color:var(--muted);font-size:13px;font-weight:700}.dot{width:10px;height:10px;border-radius:50%}.dot.ready{background:var(--green)}.dot.building{background:var(--amber)}.dot.error{background:var(--red)}.dot.idle{background:#9aa6b8}
.hero{margin-top:18px;padding:26px 28px}.hero-top{display:flex;justify-content:space-between;gap:16px;align-items:flex-start;flex-wrap:wrap}.hero h2{margin:8px 0 4px;font-size:42px;line-height:1.02}.eyebrow{font-size:12px;font-weight:800;color:var(--blue);letter-spacing:.14em;text-transform:uppercase}
.meta-grid{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:12px;margin-top:18px}.meta-pill{background:var(--panel2);border:1px solid var(--line);border-radius:18px;padding:14px 16px}.meta-pill .label{font-size:12px;color:var(--muted);text-transform:uppercase;font-weight:800;letter-spacing:.08em}.meta-pill .value{margin-top:6px;font-size:16px;font-weight:700}
.health{padding:10px 14px;border-radius:999px;font-size:13px;font-weight:800}.health.ok{background:#eaf8f0;color:var(--green)}.health.warn{background:#fff6e6;color:var(--amber)}.health.bad{background:#fff0f0;color:var(--red)}
.kpis{margin-top:18px;padding:16px}.kpi-grid{display:grid;grid-template-columns:repeat(6,minmax(0,1fr));gap:14px}.kpi{background:linear-gradient(180deg,#fff,#f8faff);border:1px solid var(--line);border-radius:20px;padding:18px;min-height:132px}.kpi .label{font-size:13px;color:var(--muted);font-weight:800;text-transform:uppercase;letter-spacing:.06em}.kpi .value{margin-top:10px;font-size:34px;font-weight:800;line-height:1}.kpi .sub{margin-top:10px;font-size:13px;color:var(--muted)}
.kpi.good{background:linear-gradient(180deg,#ffffff,#ebf9f1);border-color:#8fd5b0}.kpi.warn{background:linear-gradient(180deg,#ffffff,#fff4e1);border-color:#f1c171}.kpi.bad{background:linear-gradient(180deg,#ffffff,#ffe9e9);border-color:#ee9a9a}.prod-pilot{margin-top:14px;border:1px solid #d7e1ef;border-radius:18px;padding:16px;background:linear-gradient(180deg,#ffffff,#f7faff)}.prod-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px}.prod-card{border:1px solid #dbe4f2;border-radius:14px;padding:14px;background:#fff}.prod-card .t{font-size:12px;color:#5f6f88;font-weight:800;text-transform:uppercase;letter-spacing:.06em}.prod-card .v{margin-top:8px;font-size:30px;font-weight:900}.prod-card .s{margin-top:6px;font-size:13px;color:#6e7a90}.prod-card.future{background:linear-gradient(180deg,#fff,#f3f7ff)}
.layout{display:grid;grid-template-columns:1fr;gap:18px;margin-top:18px}.section{padding:18px 18px 22px}.section h3{margin:0 0 14px;font-size:24px}
.chart-card{min-height:560px}.chart-wrap{height:340px;border-radius:18px;background:linear-gradient(180deg,#fbfcff 0%,#f4f7fd 100%);border:1px solid #d5deec;padding:16px}.cum-wrap{margin-top:14px;padding:16px;border:1px solid #d5deec;border-radius:16px;background:linear-gradient(180deg,#ffffff 0%,#f8fbff 100%)}.cum-title{margin:0 0 10px;font-size:16px;font-weight:900}.cum-row{display:grid;grid-template-columns:200px 1fr 130px;gap:12px;align-items:center;margin:10px 0}.cum-track{height:12px;background:#e7edf7;border-radius:999px;overflow:hidden}.cum-fill{height:100%}.cum-pre{background:#a6bbe6}.cum-fac{background:#ef8d00}
.legend{display:flex;gap:18px;align-items:center;font-size:13px;color:var(--muted);font-weight:700;margin-top:10px}.legend span{display:inline-flex;align-items:center;gap:8px}.swatch{display:inline-block;width:14px;height:14px;border-radius:4px}
.table-wrap{overflow:auto;border:1px solid var(--line);border-radius:18px}table{width:100%;border-collapse:collapse}th,td{padding:14px;border-bottom:1px solid var(--line);font-size:14px;text-align:left}th{background:#f7f9fc;color:#536079;font-size:12px;text-transform:uppercase;letter-spacing:.08em}tr:last-child td{border-bottom:none}td.num{text-align:right;font-variant-numeric:tabular-nums}
.delta.pos{color:var(--green);font-weight:800}.delta.neg{color:var(--red);font-weight:800}.insights{display:flex;flex-wrap:wrap;gap:10px}.insight{padding:12px 14px;border-radius:16px;font-size:14px;font-weight:700;border:1px solid var(--line);background:var(--panel2)}
.notice{padding:14px 16px;border-radius:16px;background:#fff7e7;color:#8c6211;border:1px solid #f0dcab}.error{padding:14px 16px;border-radius:16px;background:#fff0f0;color:#992f2f;border:1px solid #f1c6c6}.empty{padding:28px;border:1px dashed var(--line);border-radius:18px;color:var(--muted);text-align:center;background:#fafbfd}.small{font-size:13px;color:var(--muted)}.footer-row{display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap;margin-top:12px}
@media (max-width:1200px){.kpi-grid{grid-template-columns:repeat(2,1fr)}.prod-grid{grid-template-columns:repeat(2,1fr)}.layout{grid-template-columns:1fr}.meta-grid{grid-template-columns:repeat(2,1fr)}}@media (max-width:720px){.topbar{position:static}.kpi-grid{grid-template-columns:1fr}.prod-grid{grid-template-columns:1fr}.meta-grid{grid-template-columns:1fr}.hero h2{font-size:30px}.select,.search{min-width:100%}}
</style>
</head>
<body>
<div class='container'>
  <div class='topbar'>
    <div class='brand'><div class='eyebrow'>Gestion affaire</div><h1>Finance</h1><p>Cockpit financier par affaire</p></div>
    <div class='controls'>
      <input id='searchInput' class='search locked' type='search' placeholder='Projet verrouillé (changer depuis Accueil)' readonly>
      <select id='affaireSelect' class='select locked' disabled><option value=''>Projet verrouillé</option></select>
      <button id='reloadBtn' class='btn primary'>Reconstruire le cache</button>
      <button id='exportBtn' class='btn dark' disabled>Exporter CSV</button>
      <a id='pmBtn' class='btn dark' href='/gestion-projet'>Gestion de projet</a>
      <a class='btn dark' href='/'>Accueil</a>
      <a id='dashboardBtn' class='btn dark' href='/dashboard'>Tableau de bord</a>
      <div id='cacheBadge' class='badge'><span class='dot idle'></span><span>Cache : attente</span></div>
    </div>
  </div>

  <div id='errorBox' style='display:none;margin-top:18px' class='error'></div>
  <div id='noticeBox' style='display:none;margin-top:18px' class='notice'></div>

  <div class='hero'>
    <div class='hero-top'>
      <div>
        <div class='eyebrow'>Affaire sélectionnée</div>
        <h2 id='heroTitle'>Sélectionnez une affaire</h2>
        <div id='heroSubtitle' class='small'>Le cockpit se remplit à partir du cache du tableau activité.</div>
      </div>
      <div id='heroHealth' class='health warn'>En attente</div>
    </div>
    <div class='meta-grid'>
      <div class='meta-pill'><div class='label'>CLIENT</div><div class='value' id='metaClient'>-</div></div>
      <div class='meta-pill'><div class='label'>PROJET</div><div class='value' id='metaProject'>-</div></div>
      <div class='meta-pill'><div class='label'>MISSIONS</div><div class='value' id='metaMissions'>-</div></div>
      <div class='meta-pill'><div class='label'>STATUT</div><div class='value' id='metaStatus'>🟠 Attention</div></div>
    </div>
  </div>



  <div class='kpis'><div class='kpi-grid'>
    <div class='kpi' id='kpiCommandeCard'><div class='label'>💰 Commandes achetées</div><div class='value' id='kpiCommande'>0 €</div><div class='sub'>Montant contractualisé</div></div>
    <div class='kpi' id='kpiAnterioriteCard'><div class='label'>📚 Antériorité</div><div class='value' id='kpiAnteriorite'>0 €</div><div class='sub'>Somme G → M</div></div>
    <div class='kpi' id='kpiFacture2026Card'><div class='label'>📈 Facturé 2026</div><div class='value' id='kpiFacture2026'>0 €</div><div class='sub'>Colonne N</div></div>
    <div class='kpi' id='kpiFacturationTotaleCard'><div class='label'>🧾 Facturation totale</div><div class='value' id='kpiFacturationTotale'>0 €</div><div class='sub'>📚 Antériorité + 2026</div></div>
    <div class='kpi' id='kpiResteCard'><div class='label'>⚠ Reste à facturer</div><div class='value' id='kpiReste'>0 €</div><div class='sub'>Solde estimé</div></div>
    <div class='kpi' id='kpiAvanceCard'><div class='label'>✅ Avancement financier</div><div class='value' id='kpiAvance'>0 %</div><div class='sub'>Facturé / commande</div></div>
  </div></div>

  <div class='prod-pilot'><h3 style='margin:0 0 12px'>Pilotage production & rentabilité (préparation)</h3><div class='prod-grid'><div class='prod-card'><div class='t'>Avancement financier</div><div id='prodFinPct' class='v'>0 %</div><div class='s'>Facturation cumulée / commande</div></div><div class='prod-card'><div class='t'>Avancement production pointé</div><div id='prodPointagePct' class='v'>0 %</div><div class='s'>Pointage opérationnel (CST)</div></div><div class='prod-card'><div class='t'>Écart financier vs production</div><div id='prodGapPct' class='v'>0 pt</div><div id='prodGapEur' class='s'>0 €</div></div><div class='prod-card future'><div class='t'>Imputation & rentabilité</div><div class='v'>À venir</div><div class='s'>Tuile prête pour l'intégration des pointages ressources</div></div></div></div>

  <div class='layout'>
    <div class='section chart-card'>
      <h3>Facturation mensuelle 2026</h3>
      <div class='chart-wrap'><svg id='monthlyChart' width='100%' height='100%' viewBox='0 0 980 320' preserveAspectRatio='none'></svg></div>
      <div class='legend'><span><i class='swatch' style='background:#dbe6ff'></i>Prévisionnel</span><span><i class='swatch' style='background:#ef8d00'></i>Facturation</span></div><div class='cum-wrap'><div class='cum-title'>Graphique cumulatif</div><div id='cumulativeChart'></div></div>
    </div>
  </div>

  <div class='section' style='margin-top:12px'><h3>Détail des missions <span id='missionsMeta' class='small' style='margin-left:8px'>0 mission</span></h3><div id='missionsTableWrap' class='table-wrap'><div class='empty'>Sélectionnez une affaire pour afficher les missions.</div></div></div>

  <div class='section' style='margin-top:12px'><h3>Insights / alertes</h3><div id='insightsBox' class='insights'><div class='empty' style='width:100%'>Sélectionnez une affaire.</div></div><div class='footer-row'><div class='small' id='statusMeta'>Cache en attente.</div></div></div>

  <div class='section' style='margin-top:18px'><h3>Mensuel détaillé</h3><div id='monthlyTableWrap' class='table-wrap'><div class='empty'>Sélectionnez une affaire pour afficher le détail mensuel.</div></div></div>

</div>

<script>
const MONTHS=["janvier","fevrier","mars","avril","mai","juin","juillet","aout","septembre","octobre","novembre","decembre"];
const MONTH_LABELS={"janvier":"Janv.","fevrier":"Févr.","mars":"Mars","avril":"Avr.","mai":"Mai","juin":"Juin","juillet":"Juil.","aout":"Août","septembre":"Sept.","octobre":"Oct.","novembre":"Nov.","decembre":"Déc."};
const state={cacheStatus:null,affaires:[],selectedAffaireId:"",selectedAffaire:null,monthlyExpanded:false};
function euro(v){return new Intl.NumberFormat('fr-FR',{style:'currency',currency:'EUR',maximumFractionDigits:0}).format(Number(v||0));}
function pct(v){return new Intl.NumberFormat('fr-FR',{style:'percent',maximumFractionDigits:1}).format(Number(v||0));}
function fmt(v){return new Intl.NumberFormat('fr-FR',{maximumFractionDigits:0}).format(Number(v||0));}
function esc(v){return String(v??'').replace(/[&<>]/g,ch=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[ch]||ch));}
function showError(msg){const b=document.getElementById('errorBox');b.textContent=msg||'Erreur';b.style.display='block';}
function clearError(){document.getElementById('errorBox').style.display='none';}
function showNotice(msg){const b=document.getElementById('noticeBox');b.textContent=msg||'';b.style.display=msg?'block':'none';}
function setCacheBadge(status,label){const c=status==='ready'?'ready':status==='building'?'building':status==='error'?'error':'idle';document.getElementById('cacheBadge').innerHTML=`<span class="dot ${c}"></span><span>${esc(label)}</span>`;}
async function api(url,options){const r=await fetch(url,options||{});const data=await r.json().catch(()=>({}));if(!r.ok) throw new Error(data.error||data.detail||data.message||`HTTP ${r.status}`);return data;}
function healthClass(a){const reste=Number(a.reste_a_facturer||0),taux=Number(a.taux_avancement_financier||0);if(taux>=0.9)return{label:'Presque soldée',cls:'ok'};if(reste>0&&taux<0.35)return{label:'À surveiller',cls:'warn'};if(reste<0)return{label:'Incohérence à vérifier',cls:'bad'};return{label:'Stable',cls:'ok'};}
async function loadCacheStatus(){const d=await api('/api/finance/cache-status');state.cacheStatus=d;const label=d.status==='ready'?`Cache prêt · ${d.affaires_count} affaires`:d.status==='building'?'Cache en reconstruction…':`Cache : ${d.status}`;setCacheBadge(d.status,label);document.getElementById('statusMeta').textContent=`${d.affaires_count||0} affaires · ${d.rows_kept||0} lignes utiles · ${d.generated_at||'pas encore généré'}`;}
async function loadAffairesList(search=''){const d=await api(`/api/finance/affaires?search=${encodeURIComponent(search)}`);state.affaires=d.items||[];const sel=document.getElementById('affaireSelect');const prev=state.selectedAffaireId;sel.innerHTML=`<option value=''>Sélectionnez une affaire</option>`+state.affaires.map(x=>`<option value="${esc(x.affaire_id)}">${esc(x.display_name)}</option>`).join('');if(prev&&state.affaires.some(x=>x.affaire_id===prev)){sel.value=prev;}else{state.selectedAffaireId='';state.selectedAffaire=null;}showNotice(state.affaires.length?`${state.affaires.length} affaire(s) disponible(s)`:'Aucune affaire trouvée pour ce filtre.');}
async function loadSelectedAffaire(id){if(!id){state.selectedAffaireId='';state.selectedAffaire=null;renderAll();return;}const d=await api(`/api/finance/affaire/${encodeURIComponent(id)}`);state.selectedAffaireId=id;state.selectedAffaire=d.affaire||null;renderAll();localStorage.setItem('selectedAffaireId',id);}
function setHeroEmpty(){document.getElementById('heroTitle').textContent='Sélectionnez une affaire';document.getElementById('heroSubtitle').textContent='Le cockpit se remplit à partir du cache du tableau activité.';document.getElementById('metaClient').textContent='-';document.getElementById('metaProject').textContent='-';document.getElementById('metaMissions').textContent='-';document.getElementById('metaStatus').textContent='🟠 Attention';const h=document.getElementById('heroHealth');h.textContent='En attente';h.className='health warn';}
function cardTone(id,tone){const el=document.getElementById(id);el.classList.remove('good','warn','bad');if(tone)el.classList.add(tone);}
function renderHero(){const a=state.selectedAffaire;if(!a){setHeroEmpty();return;}document.getElementById('heroTitle').textContent=a.display_name||'-';document.getElementById('heroSubtitle').textContent=`Client ${a.client||'-'} · ${(a.missions||[]).length} mission(s)`;document.getElementById('metaClient').textContent=a.client||'-';document.getElementById('metaProject').textContent=a.affaire||'-';document.getElementById('metaMissions').textContent=String((a.missions||[]).length||0);const hh=healthClass(a),el=document.getElementById('heroHealth');el.textContent=hh.label;el.className=`health ${hh.cls}`;document.getElementById('metaStatus').textContent=hh.cls==='ok'?'🟢 Stable':(hh.cls==='warn'?'🟠 Attention':'🔴 Risque');}
function renderKpis(){const a=state.selectedAffaire||{commande_ht:0,anteriorite:0,facture_2026:0,facturation_totale:0,reste_a_facturer:0,taux_avancement_financier:0};
document.getElementById('kpiCommande').textContent=euro(a.commande_ht);
document.getElementById('kpiAnteriorite').textContent=euro(a.anteriorite||0);
document.getElementById('kpiFacture2026').textContent=euro(a.facture_2026||a.facturation_cumulee_2026||0);
document.getElementById('kpiFacturationTotale').textContent=euro(a.facturation_totale||0);const facSub=document.querySelector('#kpiFacturationTotaleCard .sub');if(facSub){facSub.textContent=`Comparé au pointage: ${euro(a.pointage_progress_amount||0)}`;}
document.getElementById('kpiReste').textContent=euro(a.reste_a_facturer);
document.getElementById('kpiAvance').textContent=pct(a.taux_avancement_financier);const avSub=document.querySelector('#kpiAvanceCard .sub');if(avSub){avSub.textContent=`Écart facturation vs pointage: ${euro(a.pointage_vs_facturation_gap||0)}`;}
cardTone('kpiCommandeCard','good');
cardTone('kpiAnterioriteCard',(a.anteriorite||0)>0?'good':'warn');
cardTone('kpiFacture2026Card',(a.facture_2026||0)>0?'good':'warn');
cardTone('kpiFacturationTotaleCard',(a.facturation_totale||0)>0?'good':'warn');
cardTone('kpiResteCard',a.reste_a_facturer<0?'bad':(a.reste_a_facturer>(a.commande_ht||0)*0.5?'warn':'good'));
cardTone('kpiAvanceCard',a.taux_avancement_financier>0.85?'good':(a.taux_avancement_financier<0.35?'warn':''));}
function renderProductionPilot(){const a=state.selectedAffaire||{};const fin=Number(a.taux_avancement_financier||0);const prod=Number(a.pointage_progress_ratio||0);const gapPct=(fin-prod)*100;const gapEur=Number(a.pointage_vs_facturation_gap||0);const e1=document.getElementById('prodFinPct');const e2=document.getElementById('prodPointagePct');const e3=document.getElementById('prodGapPct');const e4=document.getElementById('prodGapEur');if(e1)e1.textContent=pct(fin);if(e2)e2.textContent=pct(prod);if(e3)e3.textContent=`${gapPct>=0?'+':''}${gapPct.toFixed(1)} pt`;if(e4)e4.textContent=`${euro(gapEur)} · ${gapEur>=0?'facturation en avance':'production en avance'}`;}
function renderFinanceChart(){const root=document.getElementById('monthlyChart');const a=state.selectedAffaire;if(!a){root.innerHTML=`<text x="490" y="160" text-anchor="middle" fill="#6e7a90" font-size="18">Sélectionnez une affaire</text>`;return;}const rows=MONTHS.map(m=>({label:MONTH_LABELS[m],pre:Number((((a.mensuel||{})[m]||{}).previsionnel)||0),fac:Number((((a.mensuel||{})[m]||{}).facture)||0)}));const maxVal=Math.max(1,...rows.flatMap(x=>[x.pre,x.fac]));const left=56,top=16,width=880,height=250,step=width/rows.length,groupW=Math.min(66,step*0.72),barW=Math.max(10,(groupW-8)/2);let grid='',bars='',labels='';for(let i=0;i<=4;i++){const y=top+(height/4)*i,val=Math.round(maxVal*(1-i/4));grid+=`<line x1="${left}" y1="${y}" x2="${left+width}" y2="${y}" stroke="#d8e1ef" stroke-width="1"/><text x="${left-10}" y="${y+4}" text-anchor="end" fill="#6f7f97" font-size="12">${fmt(val)}</text>`;}rows.forEach((it,i)=>{const gx=left+i*step+(step-groupW)/2;const preH=(it.pre/maxVal)*height;const facH=(it.fac/maxVal)*height;const preY=top+height-preH;const facY=top+height-facH;bars+=`<rect x="${gx}" y="${preY}" width="${barW}" height="${Math.max(preH,1)}" rx="6" fill="#cfdcf6" stroke="#b5c7eb" stroke-width="1"><title>${it.label} prévisionnel: ${euro(it.pre)}</title></rect>`;bars+=`<rect x="${gx+barW+8}" y="${facY}" width="${barW}" height="${Math.max(facH,1)}" rx="6" fill="#ef8d00" stroke="#d87800" stroke-width="1"><title>${it.label} facturation: ${euro(it.fac)}</title></rect>`;labels+=`<text x="${gx+groupW/2}" y="${top+height+22}" text-anchor="middle" fill="#5f6f88" font-size="12">${it.label}</text>`;});root.innerHTML=`${grid}<line x1="${left}" y1="${top+height}" x2="${left+width}" y2="${top+height}" stroke="#b8c3d4" stroke-width="1.2"/>${bars}${labels}`;}
function renderCumulativeChart(){const root=document.getElementById('cumulativeChart');const a=state.selectedAffaire;if(!a){root.innerHTML="<div class='small'>Sélectionnez une affaire.</div>";return;}const pre=Number(a.total_previsionnel||0);const fac=Number(a.total_facture||0);const max=Math.max(1,pre,fac);const prePct=(pre/max)*100;const facPct=(fac/max)*100;const delta=fac-pre;root.innerHTML=`<div class='cum-row'><div><strong>Prévisionnel cumulé</strong><div class='small'>Base de comparaison</div></div><div class='cum-track'><div class='cum-fill cum-pre' style='width:${prePct}%'></div></div><div><strong>${euro(pre)}</strong></div></div><div class='cum-row'><div><strong>Facturation cumulée</strong><div class='small'>Écart: ${euro(delta)}</div></div><div class='cum-track'><div class='cum-fill cum-fac' style='width:${facPct}%'></div></div><div><strong>${euro(fac)}</strong></div></div>`;}
function renderMonthlyTable(){const root=document.getElementById('monthlyTableWrap');const a=state.selectedAffaire;if(!a){root.innerHTML=`<div class='empty'>Sélectionnez une affaire pour afficher le détail mensuel.</div>`;return;}const ordered=[...MONTHS];const showAll=(state.monthlyExpanded===true);const visible=showAll?ordered:ordered.slice(0,6);let rows='';visible.forEach(m=>{const pre=Number((((a.mensuel||{})[m]||{}).previsionnel)||0),fac=Number((((a.mensuel||{})[m]||{}).facture)||0),ec=fac-pre;rows+=`<tr><td>${MONTH_LABELS[m]} 2026</td><td class='num'>${euro(pre)}</td><td class='num'>${euro(fac)}</td><td class='num delta ${ec>=0?'pos':'neg'}'>${euro(ec)}</td></tr>`;});rows+=`<tr><td><strong>Total 2026</strong></td><td class='num'><strong>${euro(a.total_previsionnel||0)}</strong></td><td class='num'><strong>${euro(a.total_facture||0)}</strong></td><td class='num delta ${Number(a.ecart_previsionnel_vs_facture||0)>=0?'pos':'neg'}'><strong>${euro(a.ecart_previsionnel_vs_facture||0)}</strong></td></tr>`;const btn=ordered.length>6?`<div style='padding:8px'><button id='btnToggleMonths' class='btn' type='button'>${showAll?'Voir moins':"Voir toute l'année 2026"}</button></div>`:'';root.innerHTML=`<table><thead><tr><th>Mois</th><th class='num'>Prévisionnel</th><th class='num'>Facturé</th><th class='num'>Écart</th></tr></thead><tbody>${rows}</tbody></table>${btn}`;const t=document.getElementById('btnToggleMonths');if(t)t.addEventListener('click',()=>{state.monthlyExpanded=!state.monthlyExpanded;renderMonthlyTable();});}
function renderMissions(){const root=document.getElementById('missionsTableWrap');const meta=document.getElementById('missionsMeta');if(!root||!meta)return;const a=state.selectedAffaire;if(!a){meta.textContent='0 mission';root.innerHTML=`<div class='empty'>Sélectionnez une affaire pour afficher les missions.</div>`;return;}const missions=a.missions||[];meta.textContent=`${missions.length} mission(s)`;if(!missions.length){root.innerHTML=`<div class='empty'>Aucune mission détaillée sur cette affaire.</div>`;return;}root.innerHTML=`<table><thead><tr><th>Tag</th><th>Mission</th><th>N°</th><th class='num'>Commande</th><th class='num'>🧾 Facturation totale</th><th class='num'>Reste</th><th class='num'>Prévisionnel</th><th class='num'>Facturé</th></tr></thead><tbody>${missions.map(m=>`<tr><td>${esc(m.tag||'')}</td><td>${esc(m.label||'')}</td><td>${esc(m.numero||'')}</td><td class='num'>${euro(m.commande_ht)}</td><td class='num'>${euro(m.facturation_totale||((m.anteriorite||0)+(m.facture_2026||m.facturation_cumulee_2026||0)))}</td><td class='num'>${euro(m.reste_a_facturer)}</td><td class='num'>${euro(m.total_previsionnel)}</td><td class='num'>${euro(m.total_facture)}</td></tr>`).join('')}</tbody></table>`;}
function renderInsights(){const root=document.getElementById('insightsBox');const a=state.selectedAffaire;if(!a){root.innerHTML=`<div class='empty' style='width:100%'>Sélectionnez une affaire.</div>`;return;}const items=a.insights||[];root.innerHTML=items.map(x=>`<div class='insight'>${esc(x)}</div>`).join('');}
function renderAll(){renderHero();renderKpis();renderProductionPilot();renderFinanceChart();renderCumulativeChart();renderMonthlyTable();renderMissions();renderInsights();document.getElementById('exportBtn').disabled=!state.selectedAffaireId;const dash=document.getElementById('dashboardBtn');dash.href=state.selectedAffaireId?`/dashboard?affaire_id=${encodeURIComponent(state.selectedAffaireId)}`:'/dashboard';const pm=document.getElementById('pmBtn');pm.href=state.selectedAffaireId?`/gestion-projet?affaire_id=${encodeURIComponent(state.selectedAffaireId)}`:'/gestion-projet';}
async function rebuildCache(){clearError();showNotice('Reconstruction du cache en cours…');await api('/api/finance/rebuild-cache',{method:'POST'});await loadCacheStatus();await loadAffairesList(document.getElementById('searchInput').value||'');if(state.selectedAffaireId&&state.affaires.some(x=>x.affaire_id===state.selectedAffaireId)){await loadSelectedAffaire(state.selectedAffaireId);}else{state.selectedAffaireId='';state.selectedAffaire=null;renderAll();}showNotice('Cache reconstruit avec succès.');}
async function initFinancePage(){clearError();try{await loadCacheStatus();await loadAffairesList('');const params=new URLSearchParams(window.location.search);const affairFromUrl=params.get('affaire_id');const affairFromStorage=localStorage.getItem('selectedAffaireId')||'';const preselected=affairFromUrl||affairFromStorage;if(preselected&&state.affaires.some(x=>x.affaire_id===preselected)){document.getElementById('affaireSelect').value=preselected;await loadSelectedAffaire(preselected);}else{renderAll();}}catch(err){showError(err.message||'Erreur de chargement');}
document.getElementById('reloadBtn').addEventListener('click',async()=>{try{await rebuildCache();}catch(err){showError(err.message||'Erreur de reconstruction du cache');}});
document.getElementById('exportBtn').addEventListener('click',()=>{if(!state.selectedAffaireId)return;window.location.href=`/api/finance/affaire/${encodeURIComponent(state.selectedAffaireId)}/export-csv`;});}
initFinancePage();
</script>
</body>
</html>
"""


def gestion_projet_html() -> str:
    return """<!doctype html>
<html lang='fr'>
<head><meta charset='utf-8'><meta name='viewport' content='width=device-width, initial-scale=1'>
<title>Gestion de projet</title>
<style>
:root{--line:#dfe5ef;--ink:#122033;--muted:#6e7a90;--accent:#ef8d00;--panel:#fff;--shadow:0 12px 34px rgba(18,32,51,.07)}
*{box-sizing:border-box}body{margin:0;font-family:Inter,Segoe UI,Arial,sans-serif;background:#f3f6fb;color:var(--ink)}
.wrap{max-width:1480px;margin:20px auto;padding:0 16px}.top,.kpis,.section{background:var(--panel);border:1px solid var(--line);border-radius:22px;box-shadow:var(--shadow)}
.top{padding:14px;display:flex;gap:10px;align-items:center;flex-wrap:wrap}.search,.select{height:44px;border:1px solid var(--line);border-radius:12px;padding:0 12px;min-width:280px}.search.locked,.select.locked{background:#eef2f8;color:#6e7a90;pointer-events:none}
.btn{height:44px;border-radius:12px;border:none;padding:0 14px;background:var(--accent);color:#fff;text-decoration:none;display:inline-flex;align-items:center;font-weight:800;cursor:pointer}
.kpis{margin-top:12px;padding:14px}.grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px}.k{border:1px solid var(--line);border-radius:14px;padding:14px;background:#fbfdff}.k .v{font-size:34px;font-weight:900;margin-top:6px}
.section{margin-top:12px;padding:14px}.subgrid{display:grid;grid-template-columns:1fr 1fr;gap:12px}.table-wrap{overflow:auto;border:1px solid var(--line);border-radius:14px}table{width:100%;border-collapse:collapse}th,td{padding:10px 12px;border-bottom:1px solid var(--line);font-size:13px;text-align:left}th{background:#f7f9fc;font-size:12px;color:#5b6880;text-transform:uppercase}.small{color:var(--muted);font-size:13px}
.bar{height:10px;background:#edf1f8;border-radius:999px;overflow:hidden}.fill{height:100%;background:#ef8d00}
.loading-wrap{margin-top:8px;padding:8px 10px;border:1px solid #d8e0ee;border-radius:12px;background:#fff}.loading-label{font-size:12px;color:#5b6880;margin-bottom:6px;font-weight:700}.loading-track{height:8px;border-radius:999px;background:#eef2f8;overflow:hidden}.loading-bar{height:100%;width:35%;background:linear-gradient(90deg,#ef8d00,#ffd08a);animation:loadmove 1.2s infinite ease-in-out}@keyframes loadmove{0%{margin-left:-35%}100%{margin-left:100%}}.match-box{margin-top:10px;border:1px solid var(--line);border-radius:10px;padding:8px;background:#fffaf2}.match-box.ok{border-color:#49a66a;background:#f4fcf6}.match-box.warn{border-color:#ef8d00;background:#fff7ec}.match-title{font-weight:700;margin-bottom:4px;font-size:13px}.conf-badge{display:inline-block;margin-left:8px;padding:1px 8px;border-radius:999px;font-size:11px;font-weight:800;background:#eef2f8;color:#3b4f6f}.match-grid{display:none;grid-template-columns:repeat(3,minmax(0,1fr));gap:8px}.match-box:hover .match-grid,.match-box:focus-within .match-grid{display:grid}.match-item{font-size:12px;color:#30425f}.match-item b{display:block;color:#6e7a90;font-size:11px;margin-bottom:2px}.mono{font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace;word-break:break-word}.pilot-box{margin-top:12px;border:1px solid #cfd8e8;border-radius:18px;padding:16px;background:#f8fbff}.pilot-title{font-size:34px;font-weight:900;margin:4px 0 10px}.pilot-grid{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:14px}.pilot-card{background:#fff;border:1px solid #d8e0ee;border-radius:18px;padding:14px}.pilot-card .t{font-size:15px;color:#4b5d7a;font-weight:800}.pilot-card .v{font-size:44px;font-weight:900;margin-top:8px}.pilot-list{margin-top:12px;border:1px solid #d8e0ee;border-radius:18px;background:#fff;padding:10px}.pilot-row{display:flex;justify-content:space-between;gap:12px;padding:12px;border:1px solid #e6ecf6;border-radius:12px;margin:8px 0;background:linear-gradient(180deg,#ffffff 0%,#f8fbff 100%)}.pilot-row:last-child{border-bottom:1px solid #e6ecf6}.pilot-row .name{font-weight:800}.pilot-row-btn{appearance:none;-webkit-appearance:none;width:100%;text-align:left;background:none;cursor:pointer;color:inherit;font:inherit}.company-cell{display:flex;align-items:center;gap:10px}.company-logo{width:28px;height:28px;border-radius:50%;border:1px solid #d8e0ee;background:#fff;object-fit:cover;display:inline-flex;align-items:center;justify-content:center;font-size:11px;color:#6e7a90;font-weight:800}.proj-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}.progress-card{border:1px solid #d8e0ee;border-radius:14px;padding:12px;background:#fff}.progress-title{font-weight:800;margin-bottom:8px}.progress-row{display:grid;grid-template-columns:180px 1fr 70px;gap:10px;align-items:center;padding:6px 0}.progress-track{height:10px;border-radius:999px;background:#edf1f8;overflow:hidden}.progress-fill{height:100%;background:#1b6ef3}.curve-wrap{border:1px solid #d8e0ee;border-radius:14px;padding:12px;background:#fff}.project-head{display:flex;align-items:center;gap:14px}.project-cover{width:88px;height:64px;border-radius:10px;border:1px solid #d8e0ee;object-fit:cover;background:#fff}.project-title{font-size:30px;font-weight:900;margin:4px 0 2px}.project-desc{color:#6e7a90;margin:2px 0 0}.timeline-box{margin:10px 0;border:1px solid #d8e0ee;border-radius:14px;background:#fff;padding:10px}.timeline-row{display:flex;justify-content:space-between;align-items:center;gap:8px;margin-bottom:6px}.timeline-track{height:12px;border-radius:999px;background:#edf1f8;overflow:hidden}.timeline-fill{height:100%;background:#1b6ef3}.timeline-fill.over{background:#d64545}.chart-controls{display:flex;gap:8px;align-items:center;margin-bottom:8px}.chart-select{height:36px;border:1px solid var(--line);border-radius:10px;padding:0 10px}.react-row{display:flex;justify-content:space-between;gap:12px;padding:8px 0;border-bottom:1px dashed #d8e0ee}.react-row:last-child{border-bottom:none}.pm-modal{position:fixed;inset:0;background:rgba(11,20,33,.45);display:flex;align-items:center;justify-content:center;padding:20px;z-index:50}.pm-modal-box{background:#fff;border-radius:14px;max-width:1100px;width:100%;max-height:82vh;overflow:auto;padding:14px;border:1px solid #d8e0ee}.lvl-row{background:#f7f9fd;font-weight:800}.task-late{background:#fff1f1}.task-soon{background:#fffaf0}
@media (max-width:980px){.grid{grid-template-columns:repeat(2,1fr)}.subgrid{grid-template-columns:1fr}}@media (max-width:640px){.grid{grid-template-columns:1fr}}
</style></head>
<body><div class='wrap'>
  <div class='top'>
    <a class='btn' href='/'>Accueil</a>
    <input id='searchInput' class='search locked' type='search' placeholder='Projet verrouillé (changer depuis Accueil)' readonly>
    <select id='affaireSelect' class='select locked' disabled><option value=''>Projet verrouillé</option></select>
    <a id='financeBtn' class='btn' href='/finance'>Finances</a>
    <a id='dashboardBtn' class='btn' href='/dashboard'>Tableau de bord</a>
    <a class='btn' href='/'>Changer de projet</a>
  </div>
  <div id='loadingWrap' class='loading-wrap' style='display:none'><div id='loadingLabel' class='loading-label'>Chargement des indicateurs projet…</div><div class='loading-track'><div class='loading-bar'></div></div></div>

  <div class='kpis'>
    <div class='project-head'><img id='projectImage' class='project-cover' alt='Projet' src=''><div><div id='projectTitle' class='project-title'>-</div><div id='projectDesc' class='project-desc'>-</div></div></div>
    <div id='timelineBox' class='timeline-box'><div class='small'>Période projet indisponible</div></div>
    <div class='grid'>
      <div class='k'><div class='small'>Sujets ouverts</div><div id='kOpen' class='v'>0</div></div>
      <div class='k'><div class='small'>Sujets en retard</div><div id='kLate' class='v'>0</div></div>
      <div class='k'><div class='small'>Rappels ouverts (>=7j ouvrés)</div><div id='kTopRappels' class='v'>0</div></div>
      <div class='k'><div class='small'>À suivre ouverts</div><div id='kTopASuivre' class='v'>0</div></div>
    </div>
    <div id='matchBox' class='match-box warn' tabindex='0'>
      <div id='matchStatus' class='match-title'>Diagnostic de matching METRONOME</div>
      <div id='matchReason' class='small'>Aucune affaire sélectionnée.</div>
      <div class='match-grid'>
        <div id='matchSearchName' class='match-item'><b>Projet recherché</b>-</div>
        <div id='matchSearchSlug' class='match-item'><b>Slug recherché</b><span class='mono'>-</span></div>
        <div id='matchFoundName' class='match-item'><b>Projet matché</b>-</div>
        <div id='matchFoundSlug' class='match-item'><b>Slug matché</b><span class='mono'>-</span></div>
        <div id='matchResolvedTitle' class='match-item'><b>Projet résolu (title)</b>-</div>
        <div id='matchResolutionMode' class='match-item'><b>Mode de résolution</b>-</div>
        <div id='matchScore' class='match-item'><b>Score de match</b>-</div>
        <div id='rowsByTitle' class='match-item'><b>Lignes filtrées par titre</b>-</div>
        <div id='rowsById' class='match-item'><b>Lignes filtrées par ID</b>-</div>
        <div id='missingFiles' class='match-item'><b>Fichiers CSV manquants</b>-</div>
      </div>
    </div>
  </div>

  <div class='section'>
    <div class='pilot-box'>
      <h3 style='margin:0 0 10px'>KPI réunion sélectionnée</h3>
      <div class='pilot-grid'>
        <div class='pilot-card'><div class='t'>Rappels ouverts à date</div><div id='kRappelsDate' class='v'>0</div></div>
        <div class='pilot-card'><div class='t'>À suivre ouverts</div><div id='kASuivre' class='v'>0</div></div>
        <div class='pilot-card'><div class='t'>Date de référence</div><div id='kDateRef' class='v' style='font-size:34px'>-</div></div>
      </div>
      <h3 style='margin:14px 0 8px'>Rappels ouverts à date cumulés par entreprise</h3>
      <div id='pilotByCompany' class='pilot-list'><div class='small'>Aucune donnée</div></div>
    </div>
  </div>

  <div class='section'>
    <h3>Graphique par entreprise</h3>
    <div class='chart-controls'>
      <label for='companyMetric' class='small'>Afficher :</label>
      <select id='companyMetric' class='chart-select'>
        <option value='open'>Tâches ouvertes par entreprise</option>
        <option value='followup'>Tâches à suivre par entreprise</option>
        <option value='reminder'>Total de taches en rappel par entreprise</option>
      </select>
    </div>
    <div id='companyChartList' class='pilot-list'><div class='small'>Aucune donnée</div></div>
  </div>

  <div class='section'>
    <h3>Réactivité par entreprise (écart moyen échéance → clôture)</h3>
    <div id='reactivityList' class='pilot-list'><div class='small'>Aucune donnée</div></div>
  </div>

  <div class='section'>
    <h3>Pointage opérationnel</h3>
    <div class='chart-controls'>
      <input id='planningCsvInput' type='file' accept='.csv' style='display:none'>
      <input id='suiviImportInput' type='file' accept='.json' style='display:none'>
      <button id='btnImportPlanning' class='btn'>Importer planning CSV</button>
      <button id='btnExportSuivi' class='btn'>Exporter suivi (.json)</button>
      <button id='btnImportSuivi' class='btn'>Importer suivi</button>
      <input id='cetMembersInput' class='search' style='height:36px;min-width:320px' placeholder='Équipe CET (ex: CVC, PLB, ELE)'>
      <button id='btnExpandAll' class='btn'>Tout déployer</button>
      <button id='btnCollapseAll' class='btn'>Tout refermer</button>
    </div>
    <div id='pointageWrap' class='table-wrap'><div class='small' style='padding:10px'>Aucun planning importé.</div></div>
  </div>

  <div id='reminderModal' class='pm-modal' style='display:none'>
    <div class='pm-modal-box'>
      <div style='display:flex;justify-content:space-between;align-items:center'>
        <h3 id='reminderModalTitle' style='margin:0'>Rappels ouverts</h3>
        <button id='reminderModalClose' class='btn'>Fermer</button>
      </div>
      <div id='reminderModalBody' class='table-wrap' style='margin-top:8px'></div>
    </div>
  </div>

</div>
<script>
const state={affaires:[],selectedId:'',selectedLabel:'',board:null,pointage:null};
function resolveAffaire(input){const raw=String(input||'').trim();if(!raw)return null;const low=raw.toLowerCase();const slug=raw.toLowerCase().normalize('NFD').replace(/[̀-ͯ]/g,'').replace(/[^a-z0-9]+/g,'-').replace(/^-+|-+$/g,'');return (state.affaires||[]).find(a=>String(a.affaire_id||'')===raw)|| (state.affaires||[]).find(a=>String(a.display_name||'').toLowerCase()===low)|| (state.affaires||[]).find(a=>String(a.display_name||'').toLowerCase().startsWith(low))|| (state.affaires||[]).find(a=>String((a.display_name||'').toLowerCase().normalize('NFD').replace(/[̀-ͯ]/g,'').replace(/[^a-z0-9]+/g,'-').replace(/^-+|-+$/g,''))===slug)||null;}
function esc(v){return String(v??'').replace(/[&<>]/g,ch=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[ch]||ch));}
async function api(u){const r=await fetch(u);const d=await r.json();if(!r.ok) throw new Error(d.detail||'Erreur API');return d;}
function renderBars(id,items){const root=document.getElementById(id);if(!root)return;if(!items||!items.length){root.innerHTML="<div class='small'>Aucune donnée</div>";return;}const max=Math.max(1,...items.map(x=>Number(x.count||0)));root.innerHTML=items.map(x=>`<div style='margin:8px 0'><div style='display:flex;justify-content:space-between;gap:8px'><span>${esc(x.label)}</span><strong>${x.count}</strong></div><div class='bar'><div class='fill' style='width:${(Number(x.count||0)/max)*100}%'></div></div></div>`).join('');}
function renderTable(rows){const root=document.getElementById('boardTable');if(!root)return;if(!rows||!rows.length){root.innerHTML="<div class='small' style='padding:12px'>Aucun sujet ouvert.</div>";return;}root.innerHTML=`<table><thead><tr><th>Zone</th><th>Lot</th><th>Sujet</th><th>Entreprise</th><th>Responsable</th><th>Statut</th><th>Date échéance</th><th>Réunion origine</th><th>Commentaire</th></tr></thead><tbody>${rows.map(r=>`<tr><td>${esc(r.zone)}</td><td>${esc(r.lot)}</td><td>${esc(r.sujet)}</td><td>${esc(r.entreprise)}</td><td>${esc(r.responsable)}</td><td>${esc(r.statut)}</td><td>${esc(r.date_echeance)}</td><td>${esc(r.reunion_origine)}</td><td>${esc(r.commentaire)}</td></tr>`).join('')}</tbody></table>`;}
function setText(id,value){const el=document.getElementById(id);if(el) el.textContent=value;}
function setHtml(id,value){const el=document.getElementById(id);if(el) el.innerHTML=value;}
function showLoading(on,label='Chargement des indicateurs projet…'){const w=document.getElementById('loadingWrap');if(!w) return;w.style.display=on?'block':'none';setText('loadingLabel',label);}
function fmtDateFr(v){if(!v)return '-';const d=new Date(v+'T00:00:00');if(Number.isNaN(d.getTime()))return v;return d.toLocaleDateString('fr-FR',{day:'numeric',month:'long',year:'numeric'});} 
function fmtDateFrShort(v){const d=parseDateOnly(v);if(!d)return '';const yy=String(d.getFullYear()).slice(-2);const mm=String(d.getMonth()+1).padStart(2,'0');const dd=String(d.getDate()).padStart(2,'0');return `${dd}/${mm}/${yy}`;}
function parseDateOnly(v){const raw=String(v||'').trim();if(!raw)return null;const s=raw.split(' ')[0].split('T')[0];if(!s)return null;if(/^\d{4}-\d{2}-\d{2}$/.test(s)){const d=new Date(s+'T00:00:00');return Number.isNaN(d.getTime())?null:d;}const m=s.match(/^(\d{2})\/(\d{2})\/(\d{2,4})$/);if(m){const y=m[3].length===2?Number(`20${m[3]}`):Number(m[3]);const d=new Date(`${y}-${m[2]}-${m[1]}T00:00:00`);return Number.isNaN(d.getTime())?null:d;}return null;}
function reminderWeeks(deadline){const d=parseDateOnly(deadline);if(!d)return 0;const now=new Date();const today=new Date(now.getFullYear(),now.getMonth(),now.getDate());const diff=Math.floor((today-d)/(1000*60*60*24));if(diff<=0)return 0;return Math.floor(diff/7)+1;}
function renderProjectTimeline(b){const p=((b&&b.analytics)||{}).kpi_project_progress||{};const root=document.getElementById('timelineBox');if(!root)return;const total=Number(p.total_days||0);if(!total){root.innerHTML="<div class='small'>Période projet indisponible</div>";return;}const elapsedRaw=Number(p.elapsed_days_raw||0);const over=Number(p.overrun_days||0);const overPct=Number(p.overrun_pct||0);const basePct=Math.max(0,Math.min(100,Number(p.calendar_progress_pct||0)));const fillClass=(p.is_overrun?'timeline-fill over':'timeline-fill');const extra= p.is_overrun ? ` · Dépassement: <strong>${over} jours (${overPct}%)</strong>` : '';root.innerHTML=`<div class='timeline-row'><span><strong>Progression temporelle</strong></span><span><strong>${basePct.toFixed(0)}%</strong></span></div><div class='timeline-track'><div class='${fillClass}' style='width:${basePct}%'></div></div><div class='small' style='margin-top:6px'>Commencé le ${esc(fmtDateFr(p.start_date))}.<br>Fin prévue le ${esc(fmtDateFr(p.end_date))}.<br>${elapsedRaw}/${total} jours écoulés${extra}</div>`;}
function renderCompanyChart(b){const p=(b&&b.kpis_pilotage)||{};const metricEl=document.getElementById('companyMetric');const metric=(metricEl&&metricEl.value)||'open';const views=p.project_company_views||{};const items=views[metric]||[];const root=document.getElementById('companyChartList');if(!root)return;if(!items.length){root.innerHTML="<div class='small'>Aucune donnée</div>";return;}const max=Math.max(1,...items.map(x=>Number(x.count||0)));root.innerHTML=items.map(x=>{const val=Number(x.count||0);return `<div class='pilot-row' style='display:block;border:1px solid #e7edf7;border-radius:12px;margin:8px 0;padding:10px;background:#fff'><div style='display:flex;justify-content:space-between;align-items:center;gap:8px'><span class='company-cell'>${companyLogoHtml(x)}<span class='name'>${esc(x.name)}</span></span><strong>${val}</strong></div><div class='bar' style='margin-top:6px'><div class='fill' style='width:${(val/max)*100}%'></div></div></div>`;}).join('');}
function renderReactivity(b){const p=(b&&b.kpis_pilotage)||{};const items=p.reactivity_by_company||[];const root=document.getElementById('reactivityList');if(!root)return;if(!items.length){root.innerHTML="<div class='small'>Pas assez de tâches clôturées avec échéance.</div>";return;}const vals=items.map(x=>Math.abs(Number(x.avg_gap_days||0)));const max=Math.max(1,...vals);root.innerHTML=`<div style='display:flex;align-items:flex-end;gap:14px;height:240px;padding:10px 0'>${items.map(x=>{const v=Number(x.avg_gap_days||0);const h=Math.max(8,(Math.abs(v)/max)*190);const color=v>0?'#d64545':'#1b6ef3';return `<div style='flex:1;min-width:60px;text-align:center'><div title='${esc(x.name)}: ${v} j' style='margin:0 auto;width:34px;height:${h}px;background:${color};border-radius:8px 8px 0 0'></div><div style='font-size:11px;margin-top:6px;font-weight:700'>${esc(x.name)}</div><div style='font-size:11px;color:#6e7a90'>${v} j</div></div>`;}).join('')}</div>`;}
function initials(v){const parts=String(v||'').trim().split(/\s+/).slice(0,2);return parts.map(x=>x[0]||'').join('').toUpperCase()||'?';}
function companyLogoHtml(item){const name=item.name||item.label||'';const logo=item.logo||'';if(logo){return `<img class='company-logo' src='${esc(logo)}' alt='${esc(name)}'>`;}return `<span class='company-logo'>${esc(initials(name))}</span>`;}
function showReminderModal(company,b){const p=(b&&b.kpis_pilotage)||{};const norm=v=>String(v||'').toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'').trim();const items=(p.reminders_open_items||[]).filter(x=>norm(x.company)===norm(company)).filter(x=>reminderWeeks(x.deadline)>0);const body=document.getElementById('reminderModalBody');const title=document.getElementById('reminderModalTitle');if(title)title.textContent=`Rappels ouverts - ${company}`;if(!body)return;if(!items.length){body.innerHTML="<div class='small' style='padding:10px'>Aucun rappel ouvert.</div>";}else{const byPer={};for(const it of items){const k=it.perimetre||'Général';(byPer[k]=byPer[k]||[]).push(it);}let html='';for(const per of Object.keys(byPer).sort()){html+=`<h4 style='margin:10px 0 6px'>${esc(per)}</h4><table><thead><tr><th>Tâche</th><th>Échéance</th><th>Rappel</th></tr></thead><tbody>${byPer[per].map(r=>{const w=reminderWeeks(r.deadline);const warn=`<span style='color:#d64545;font-weight:800'>Rappel ${w}</span>`;return `<tr><td>${esc(r.task||'')}</td><td>${esc(fmtDateFrShort(r.deadline||''))}</td><td>${warn}</td></tr>`;}).join('')}</tbody></table>`;}body.innerHTML=html;}const modal=document.getElementById('reminderModal');if(modal)modal.style.display='flex';}
function hideReminderModal(){const modal=document.getElementById('reminderModal');if(modal)modal.style.display='none';}
async function loadBoard(id){const requested=id||state.selectedId||'';if(!requested){state.selectedId='';state.selectedLabel='';state.board=null;renderBoard();return;}const selected=resolveAffaire(requested);const aid=(selected&&selected.affaire_id)||requested;state.selectedId=aid;const sel=document.getElementById('affaireSelect');if(sel)sel.value=aid;state.selectedLabel=(selected&&selected.display_name)||state.selectedLabel||'';localStorage.setItem('selectedAffaireId',aid);showLoading(true,'Chargement de la vue projet…');try{let b;try{b=await api(`/api/project-management/board?affaire_id=${encodeURIComponent(aid)}`);}catch(errById){const msg=String((errById&&errById.message)||'');if(state.selectedLabel&&(msg.includes('Affaire introuvable')||msg.includes('affaire_id ou affaire_name requis')||msg.includes('404'))){const alt=String(state.selectedLabel||'').split(' - ').slice(-1)[0]||state.selectedLabel;b=await api(`/api/project-management/board?affaire_name=${encodeURIComponent(alt)}`);}else{throw errById;}}state.board=b;renderBoard();}catch(err){console.error(err);state.board=null;renderBoard();showPageError(err);}finally{showLoading(false);}}
function renderMatchDebug(b){const md=(b&&b.match_debug)||{};const box=document.getElementById('matchBox');const status=document.getElementById('matchStatus');const reason=document.getElementById('matchReason');if(box){box.className=`match-box ${(b&&b.warning)?'warn':'ok'}`;}if(status){const conf=(b&&b.confidence_level)||'low';status.innerHTML=`Diagnostic de matching METRONOME <span class='conf-badge'>${esc(conf)}</span>`;}if(reason){reason.textContent=(b&&b.warning_message)||((b&&b.project_name)?`Projet chargé : ${b.project_name}`:'Aucune affaire sélectionnée.');}setHtml('matchSearchName',`<b>Projet recherché</b>${esc(md.searched_project_name||'-')}`);setHtml('matchSearchSlug',`<b>Slug recherché</b><span class='mono'>${esc(md.searched_project_slug||'-')}</span>`);setHtml('matchFoundName',`<b>Projet matché</b>${esc(md.matched_project_name||'-')}`);setHtml('matchFoundSlug',`<b>Slug matché</b><span class='mono'>${esc(md.matched_project_slug||'-')}</span>`);setHtml('matchResolvedTitle',`<b>Projet résolu (title)</b>${esc(md.resolved_project_title||'-')}`);setHtml('matchResolutionMode',`<b>Mode de résolution</b>${esc(md.resolution_mode||'-')}`);setHtml('matchScore',`<b>Score de match</b>${esc(String(md.match_score??'-'))}`);setHtml('rowsByTitle',`<b>Lignes filtrées par titre</b>${esc(String(md.rows_filtered_by_title??'-'))}`);setHtml('rowsById',`<b>Lignes filtrées par ID</b>${esc(String(md.rows_filtered_by_id??'-'))}`);const miss=((b&&b.missing_files)||[]);setHtml('missingFiles',`<b>Fichiers CSV manquants</b>${esc(miss.length?miss.join(', '):'-')}`);} 
function renderBoard(){const b=state.board||{};const p=(b&&b.kpis_pilotage)||{};const k=(b&&b.kpis)||{};setText('projectTitle',b.project_name||'-');setText('projectDesc',b.project_description||'-');const img=document.getElementById('projectImage');if(img){img.src=b.project_image||'';}renderMatchDebug(b);setText('kOpen',Number(k.open_topics||0));setText('kOverdue',Number(k.overdue_topics||0));setText('kTopRappels',Number(p.reminders_open_count||0));setText('kDateRef',p.date_reference||p.reference_date||'-');setText('kRappelsDate',Number(p.rappels_ouverts_a_date||p.reminders_open_count||0));setText('kASuivre',Number(p.a_suivre_ouverts||p.followups_open_count||0));renderBars('kByStatus',k.by_status||[]);renderTable(b.rows||[]);renderProjectTimeline(b);renderCompanyChart(b);renderReactivity(b);const companyRows=(p.rappels_cumules_par_entreprise||p.reminders_by_company||[]);setHtml('pilotByCompany',companyRows.length?companyRows.map(r=>{const nm=r.name||r.company||'';return `<button class='pilot-row-btn pilot-row' data-company='${esc(nm)}'><span class='company-cell'>${companyLogoHtml({name:nm,logo:r.logo||''})}<span class='name'>${esc(nm)}</span></span><span><strong>${Number(r.count||0)}</strong> rappels</span></button>`;}).join(''):"<div class='small'>Aucune donnée</div>");Array.from(document.querySelectorAll('.pilot-row-btn')).forEach(btn=>btn.addEventListener('click',()=>showReminderModal(btn.getAttribute('data-company')||'',b)));}
function showPageError(err){const root=document.getElementById('boardTable');if(root)root.innerHTML=`<div class='small' style='padding:12px;color:#b42318'>${esc(err?.message||'Erreur de chargement')}</div>`;}
async function loadAffaires(q){const data=await api(`/api/finance/affaires?search=${encodeURIComponent(q||'')}`);state.affaires=data.items||[];const sel=document.getElementById('affaireSelect');if(!sel)return;sel.innerHTML=`<option value=''>Sélectionner une affaire…</option>${state.affaires.map(a=>`<option value='${esc(a.affaire_id)}'>${esc(a.display_name||a.affaire_id)}</option>`).join('')}`;if(state.selectedId){const selected=state.affaires.find(a=>String(a.affaire_id||'')===String(state.selectedId||''));if(selected){state.selectedLabel=selected.display_name||state.selectedLabel||'';sel.value=selected.affaire_id;}}}
function getProjectKey(){return state.selectedId?`id::${state.selectedId}`:'';}
async function pointageApi(url,opts){const r=await fetch(url,opts);const d=await r.json();if(!r.ok) throw new Error(d.detail||'Erreur pointage');return d;}
async function loadPointage(){if(!state.selectedId)return;try{const d=await pointageApi(`/api/project-management/pointage?affaire_id=${encodeURIComponent(state.selectedId)}`);state.pointage=d;renderPointage();}catch(e){console.warn(e);}}
function levelGroups(tasks){const g={};for(const t of (tasks||[])){const lvl=t.level_label||t.level||'Général';(g[lvl]=g[lvl]||[]).push(t);}return g;}
function renderPointage(){const root=document.getElementById('pointageWrap');const d=state.pointage||{};const tasks=d.tasks||[];if(!root)return;if(!tasks.length){root.innerHTML="<div class='small' style='padding:10px'>Aucun planning importé.</div>";return;}const expanded=new Set((d.workState&&d.workState.expanded_levels)||[]);const groups=levelGroups(tasks);let html=`<table><thead><tr><th>Tâche</th><th>Zone</th><th>Ressource</th><th>Début</th><th>Fin</th><th>% achevé</th><th>Réception réelle</th><th>Retard</th><th>Coût planifié CST</th><th>Coût pointé CST</th><th>Statut</th></tr></thead><tbody>`;for(const lvl of Object.keys(groups)){const open=expanded.has(lvl);html+=`<tr class='lvl-row'><td colspan='11'><button type='button' data-lvl='${esc(lvl)}' class='btn' style='height:28px'>${open?'−':'+'}</button> ${esc(lvl)}</td></tr>`;if(open){for(const t of groups[lvl]){const isSummary=Boolean(t.is_summary);const rowCls=isSummary?'lvl-row':(t.isLate?'task-late':(((t.status==='future')&&t.start)?'task-soon':''));const label=t.name;const start=fmtDateFrShort(t.start||'');const end=fmtDateFrShort(t.end||'');const depth=Math.max(0,Number(t.depth||0));const pad=12+(depth*14);if(isSummary){html+=`<tr class='${rowCls}'><td style='padding-left:${pad}px;font-weight:800'>${esc(t.name)}</td><td>${esc(t.level_label||t.level||'')}</td><td>${esc(t.owner||'')}</td><td>${esc(start)}</td><td>${esc(end)}</td><td>${Number(t.progress||0).toFixed(0)}%</td><td></td><td></td><td></td><td></td><td>summary</td></tr>`;continue;}html+=`<tr class='${rowCls}'><td style='padding-left:${pad}px'>${esc(label)}</td><td>${esc(t.level_label||t.level||'')}</td><td>${esc((t.owner==='Non attribué'?'':t.owner))}</td><td>${esc(start)}</td><td>${esc(end)}</td><td><input data-task='${esc(t.task_id)}' data-field='progress' type='number' min='0' max='100' value='${Number(t.progress||0)}' style='width:70px'></td><td><input data-task='${esc(t.task_id)}' data-field='actualEnd' type='date' value='${esc(t.actualEnd||'')}'></td><td>${Number(t.delayDays||0)} j</td><td style='color:#b45f06;font-weight:800'>${Number(t.plannedCostCst||0)>0?Number(t.plannedCostCst||0).toFixed(0)+' €':''}</td><td>${Number(t.actualCostCst||0)>0?Number(t.actualCostCst||0).toFixed(0)+' €':''}</td><td>${t.status==='past'?'<span style="color:#1f7a3f;font-weight:800">Terminé</span>':(t.status==='future'?'<span style="color:#8a94a8;font-weight:800">À venir</span>':'<span style="color:#1b6ef3;font-weight:800">En cours</span>')}</td></tr>`;if(t.is_cet&&t.cetMembers&&t.cetMembers.length){for(const m of t.cetMembers){const cm=((t.cet||{})[m]||{});html+=`<tr><td style='padding-left:${pad+16}px'>↳ ${esc(m)}</td><td>${esc(t.level||'')}</td><td>CET</td><td></td><td></td><td><input data-task='${esc(t.task_id)}' data-cet='${esc(m)}' data-field='progress' type='number' min='0' max='100' value='${Number(cm.progress||0)}' style='width:70px'></td><td><input data-task='${esc(t.task_id)}' data-cet='${esc(m)}' data-field='actualEnd' type='date' value='${esc(cm.actualEnd||'')}'></td><td colspan='4'></td></tr>`;}}}}}html+='</tbody></table>';root.innerHTML=html;Array.from(root.querySelectorAll('button[data-lvl]')).forEach(b=>b.addEventListener('click',async()=>{const lvl=b.getAttribute('data-lvl')||'';const ws={...(d.workState||{}),expanded_levels:[...new Set([...(d.workState?.expanded_levels||[])]) ]};const set=new Set(ws.expanded_levels);if(set.has(lvl))set.delete(lvl);else set.add(lvl);ws.expanded_levels=[...set];await savePointage({},ws);}));Array.from(root.querySelectorAll('input[data-task]')).forEach(inp=>inp.addEventListener('change',async()=>{const task=inp.getAttribute('data-task')||'';const field=inp.getAttribute('data-field')||'';const cet=inp.getAttribute('data-cet')||'';const patch={};if(cet){patch[task]={cet:{[cet]:{[field]:inp.value}}};}else{patch[task]={[field]:inp.value};}await savePointage(patch,d.workState||{});}));}
async function savePointage(pointagePatch,workState){if(!state.selectedId)return;const cetMembers=(document.getElementById('cetMembersInput')||{}).value||'';const payload={pointage_patch:{...pointagePatch,__cetMembers:cetMembers},work_state:workState||{}};const d=await pointageApi(`/api/project-management/pointage/save?affaire_id=${encodeURIComponent(state.selectedId)}`,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});state.pointage=d;renderPointage();}
async function importPlanningFile(file){if(!state.selectedId||!file)return;const fd=new FormData();fd.append('file',file);const d=await fetch(`/api/project-management/pointage/planning/import?affaire_id=${encodeURIComponent(state.selectedId)}`,{method:'POST',body:fd});const j=await d.json();if(!d.ok)throw new Error(j.detail||'Import planning impossible');state.pointage=j;renderPointage();}
async function importSuiviFile(file){if(!state.selectedId||!file)return;const fd=new FormData();fd.append('file',file);const d=await fetch(`/api/project-management/pointage/import-suivi?affaire_id=${encodeURIComponent(state.selectedId)}`,{method:'POST',body:fd});const j=await d.json();if(!d.ok)throw new Error(j.detail||'Import suivi impossible');state.pointage=j;renderPointage();}
async function exportSuivi(){if(!state.selectedId)return;const d=await pointageApi(`/api/project-management/pointage/export?affaire_id=${encodeURIComponent(state.selectedId)}`);const blob=new Blob([JSON.stringify(d,null,2)],{type:'application/json'});const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download=`suivi-pointage-${state.selectedId}.json`;a.click();URL.revokeObjectURL(a.href);}
async function init(){
  try{
    await loadAffaires('');
    const params=new URLSearchParams(window.location.search);
    const urlId=params.get('affaire_id');
    const initialId=urlId||localStorage.getItem('selectedAffaireId')||'';
    if(initialId){
      const sel=document.getElementById('affaireSelect');
      const found=resolveAffaire(initialId);
      state.selectedId=(found&&found.affaire_id)||initialId;
      if(found){
        state.selectedId=found.affaire_id;
        sel.value=found.affaire_id;
        state.selectedLabel=found.display_name||'';
        localStorage.setItem('selectedAffaireId',found.affaire_id);
      }
      await loadBoard(state.selectedId||initialId);
      await loadPointage();
    }else{renderBoard();}
    document.getElementById('btnImportPlanning').addEventListener('click',()=>document.getElementById('planningCsvInput').click());
    document.getElementById('planningCsvInput').addEventListener('change',async e=>{if(e.target.files&&e.target.files[0]){await importPlanningFile(e.target.files[0]);e.target.value='';}});
    document.getElementById('btnImportSuivi').addEventListener('click',()=>document.getElementById('suiviImportInput').click());
    document.getElementById('suiviImportInput').addEventListener('change',async e=>{if(e.target.files&&e.target.files[0]){await importSuiviFile(e.target.files[0]);e.target.value='';}});
    document.getElementById('btnExportSuivi').addEventListener('click',exportSuivi);
    document.getElementById('btnExpandAll').addEventListener('click',async()=>{const lvls=[...new Set((state.pointage?.tasks||[]).filter(x=>x.is_summary).map(x=>x.level_label||x.level||'Général'))];await savePointage({}, {expanded_levels:lvls});});
    document.getElementById('btnCollapseAll').addEventListener('click',async()=>{await savePointage({}, {expanded_levels:[]});});
    document.getElementById('cetMembersInput').addEventListener('change',async()=>{await savePointage({}, state.pointage?.workState||{});});
    const reminderClose=document.getElementById('reminderModalClose');if(reminderClose)reminderClose.addEventListener('click',hideReminderModal);
    const affaireSelect=document.getElementById('affaireSelect');if(affaireSelect)affaireSelect.addEventListener('change',async e=>{await loadBoard(e.target.value||'');});
    const companyMetric=document.getElementById('companyMetric');if(companyMetric)companyMetric.addEventListener('change',()=>renderCompanyChart(state.board||{}));
  }catch(err){console.error(err);renderBoard();showPageError(err);}
}
init();
</script></body></html>"""



def dashboard_html() -> str:
    return """<!doctype html>
<html lang='fr'>
<head><meta charset='utf-8'><meta name='viewport' content='width=device-width, initial-scale=1'>
<title>Gestion Affaire - Tableau de bord</title>
<style>
:root{--line:#dfe5ef;--ink:#122033;--muted:#6e7a90;--accent:#ef8d00;--panel:#fff;--shadow:0 12px 34px rgba(18,32,51,.07)}
*{box-sizing:border-box}body{margin:0;font-family:Inter,Segoe UI,Arial,sans-serif;background:#f3f6fb;color:var(--ink)}
.wrap{max-width:1320px;margin:22px auto;padding:0 16px}.top,.hero,.kpis,.chart{background:var(--panel);border:1px solid var(--line);border-radius:22px;box-shadow:var(--shadow)}
.top{padding:14px;display:flex;gap:10px;align-items:center;flex-wrap:wrap}.search,.select{height:44px;border:1px solid var(--line);border-radius:12px;padding:0 12px;min-width:280px}
.btn{height:44px;border-radius:12px;border:none;padding:0 14px;background:var(--accent);color:#fff;text-decoration:none;display:inline-flex;align-items:center;font-weight:800;cursor:pointer}
.hero{padding:18px;margin-top:14px}.eyeb{font-size:12px;font-weight:800;letter-spacing:.12em;color:var(--accent);text-transform:uppercase}.title{font-size:36px;font-weight:900;margin:6px 0}.muted{color:var(--muted)}
.kpis{padding:14px;margin-top:14px}.grid{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:12px}.k{border:1px solid var(--line);border-radius:14px;padding:14px;background:#fbfdff}.k .v{font-size:36px;font-weight:900;margin-top:6px}
.chart{padding:16px;margin-top:14px}.chart h3{margin:0 0 12px;font-size:28px}.chart-wrap{height:360px;border:1px solid #d5deec;border-radius:16px;padding:12px;background:linear-gradient(180deg,#fbfcff 0%,#f4f7fd 100%)}
@media (max-width:900px){.grid{grid-template-columns:1fr}.title{font-size:28px}}
</style></head>
<body>
<div class='wrap'>
  <div class='top'>
    <a class='btn' href='/'>Accueil</a>
    <input id='searchInput' class='search' type='search' placeholder='Rechercher une affaire'>
    <select id='affaireSelect' class='select locked' disabled><option value=''>Projet verrouillé</option></select>
    <a id='financeBtn' class='btn' href='/finance'>Finances</a>
    <a id='pmBtn' class='btn' href='/gestion-projet'>Gestion de projet</a>
    <button id='exportBtn' class='btn' disabled>Exporter CSV</button>
  </div>
  <div class='hero'>
    <div class='eyeb'>Tableau de bord</div>
    <div id='title' class='title'>Sélectionnez une affaire</div>
    <div id='subtitle' class='muted'>Synthèse principale : client, commande, antériorité, facturé 2026, facturation totale et reste à facturer.</div>
  </div>
  <div class='kpis'><div class='grid'>
    <div class='k'><div class='muted'>💰 Commandes achetées</div><div id='kCommande' class='v'>0 €</div></div>
    <div class='k'><div class='muted'>Facturation cumulée</div><div id='kFacture' class='v'>0 €</div></div>
    <div class='k'><div class='muted'>⚠ Reste à facturer</div><div id='kReste' class='v'>0 €</div></div>
  </div></div>
  <div class='chart'><h3>Prévisionnel vs facturation (année complète)</h3><div class='chart-wrap'><svg id='chart' width='100%' height='320' viewBox='0 0 980 320' preserveAspectRatio='none'></svg></div></div>
</div>
<script>
const MONTHS=['janvier','fevrier','mars','avril','mai','juin','juillet','aout','septembre','octobre','novembre','decembre'];
const MONTH_LABELS={janvier:'Janv.',fevrier:'Févr.',mars:'Mars',avril:'Avr.',mai:'Mai',juin:'Juin',juillet:'Juil.',aout:'Août',septembre:'Sept.',octobre:'Oct.',novembre:'Nov.',decembre:'Déc.'};
const state={affaires:[],selectedId:'',selected:null};
function euro(v){return new Intl.NumberFormat('fr-FR',{style:'currency',currency:'EUR',maximumFractionDigits:0}).format(Number(v||0));}
function esc(v){return String(v||'').replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]));}
async function api(u){const r=await fetch(u);const d=await r.json();if(!r.ok) throw new Error(d.detail||'Erreur API');return d;}
async function loadAffairesList(search=''){const d=await api(`/api/finance/affaires?search=${encodeURIComponent(search)}`);state.affaires=d.items||[];const sel=document.getElementById('affaireSelect');sel.innerHTML=`<option value=''>Sélectionnez une affaire</option>`+state.affaires.map(x=>`<option value="${esc(x.affaire_id)}">${esc(x.display_name)}</option>`).join('');if(state.selectedId&&state.affaires.some(x=>x.affaire_id===state.selectedId)){sel.value=state.selectedId;}}
async function loadSelectedAffaire(id){if(!id){state.selectedId='';state.selected=null;render();return;}const d=await api(`/api/finance/affaire/${encodeURIComponent(id)}`);state.selectedId=id;state.selected=d.affaire;localStorage.setItem('selectedAffaireId',id);render();}
function renderChart(){const root=document.getElementById('chart');const a=state.selected;if(!a){root.innerHTML='';return;}const data=MONTHS.map(m=>({label:MONTH_LABELS[m],pre:Number(((a.mensuel||{})[m]||{}).previsionnel||0),fac:Number(((a.mensuel||{})[m]||{}).facture||0)}));const max=Math.max(1,...data.flatMap(x=>[x.pre,x.fac]));const left=56,top=20,width=880,height=250,step=width/Math.max(1,data.length),bw=Math.min(60,step*0.5);let out='';for(let i=0;i<=4;i++){const y=top+(height/4)*i;out+=`<line x1="${left}" y1="${y}" x2="${left+width}" y2="${y}" stroke="#dde5f2"/><text x="${left-8}" y="${y+4}" text-anchor="end" fill="#6b7c98" font-size="12">${Math.round(max*(1-i/4))}</text>`;}const pts=[];data.forEach((d,i)=>{const x=left+i*step+(step-bw)/2;const h=(d.pre/max)*height;const y=top+height-h;const py=top+height-(d.fac/max)*height;out+=`<rect x="${x}" y="${y}" width="${bw}" height="${Math.max(h,1)}" rx="8" fill="#f5d6a8"><title>${d.label} prévisionnel: ${d.pre.toFixed(0)} €</title></rect><text x="${x+bw/2}" y="${top+height+18}" text-anchor="middle" fill="#5f6f88" font-size="12">${d.label}</text>`;pts.push(`${x+bw/2},${py}`);});out+=`<polyline points="${pts.join(' ')}" fill="none" stroke="#ef8d00" stroke-width="4"/>`;root.innerHTML=out;}
function render(){const a=state.selected;document.getElementById('financeBtn').href=state.selectedId?`/finance?affaire_id=${encodeURIComponent(state.selectedId)}`:'/finance';document.getElementById('pmBtn').href=state.selectedId?`/gestion-projet?affaire_id=${encodeURIComponent(state.selectedId)}`:'/gestion-projet';document.getElementById('exportBtn').disabled=!state.selectedId;document.getElementById('title').textContent=a?(a.display_name||'-'):'Sélectionnez une affaire';document.getElementById('subtitle').textContent=a?`Client ${a.client||'-'} · ${a.affaire||'-'}`:'Synthèse principale : client, commande, antériorité, facturé 2026, facturation totale et reste à facturer.';document.getElementById('kCommande').textContent=euro(a?a.commande_ht:0);document.getElementById('kFacture').textContent=euro(a?(a.facturation_totale||((a.anteriorite||0)+(a.facture_2026||a.facturation_cumulee_2026||0))):0);document.getElementById('kReste').textContent=euro(a?a.reste_a_facturer:0);renderChart();}
async function init(){await loadAffairesList('');const params=new URLSearchParams(window.location.search);const pre=params.get('affaire_id')||localStorage.getItem('selectedAffaireId')||'';if(pre&&state.affaires.some(x=>x.affaire_id===pre)){document.getElementById('affaireSelect').value=pre;await loadSelectedAffaire(pre);}else{render();}document.getElementById('searchInput').addEventListener('input',async e=>{await loadAffairesList(e.target.value||'');});document.getElementById('affaireSelect').addEventListener('change',async e=>loadSelectedAffaire(e.target.value||''));document.getElementById('exportBtn').addEventListener('click',()=>{if(state.selectedId)window.location.href=`/api/finance/affaire/${encodeURIComponent(state.selectedId)}/export-csv`;});}
init();
</script>
</body>
</html>"""



@app.get("/assets/logo-tempo")
def tempo_logo():
    if os.path.exists(TEMPO_LOGO_PATH):
        return FileResponse(TEMPO_LOGO_PATH)
    raise HTTPException(status_code=404, detail="Logo Tempo introuvable")


@app.get("/", response_class=HTMLResponse)
def landing_page():
    return landing_html()


@app.get("/finance", response_class=HTMLResponse)
def finance_page():
    return finance_html()


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard_page():
    return dashboard_html()


@app.get("/gestion-projet", response_class=HTMLResponse)
def gestion_projet_page():
    return gestion_projet_html()


@app.get("/health", response_class=JSONResponse)
def health():
    return {
        "status": "ok",
        "service": "finance",
        "workbook_exists": service.workbook_exists(),
        "workbook_path": service.workbook_path,
        "sheet_name": service.sheet_name,
    }


@app.get("/api/finance", response_class=JSONResponse)
def finance_index():
    return {
        "service": "finance",
        "status": "ok",
        "endpoints": {
            "landing": "/",
            "finance_ui": "/finance",
            "dashboard_ui": "/dashboard",
            "project_management_ui": "/gestion-projet",
            "cache_status": "/api/finance/cache-status",
            "rebuild_cache": "/api/finance/rebuild-cache",
            "affaires": "/api/finance/affaires?search=...",
            "affaire": "/api/finance/affaire/{affaire_id}",
            "affaire_export_csv": "/api/finance/affaire/{affaire_id}/export-csv",
            "health": "/health",
        },
    }


@app.get("/api/finance/cache-status", response_class=JSONResponse)
def api_cache_status():
    return service.cache_status()


@app.post("/api/finance/rebuild-cache", response_class=JSONResponse)
@app.get("/api/finance/rebuild-cache", response_class=JSONResponse)
def api_rebuild_cache():
    cache = service.rebuild_finance_cache()
    return {
        "status": "rebuilt",
        "generated_at": cache.get("generated_at"),
        "rows_read": cache.get("rows_read"),
        "rows_kept": cache.get("rows_kept"),
        "affaires_count": cache.get("affaires_count"),
        "warnings": cache.get("warnings", []),
    }


@app.get("/api/finance/affaires", response_class=JSONResponse)
def api_affaires(search: str = Query(default="")):
    cache = service.get_finance_cache()
    items = FinanceService.lightweight_affaires(cache, search=search)
    return {"ok": True, "count": len(items), "items": items}


@app.get("/api/finance/affaire/{affaire_id}", response_class=JSONResponse)
def api_affaire_detail(affaire_id: str):
    cache = service.get_finance_cache()
    item = cache.get("items", {}).get(affaire_id)
    if not item:
        raise HTTPException(status_code=404, detail=f"Affaire introuvable : {affaire_id}")
    affaire = dict(item)
    affaire.update(pointage_finance_summary(affaire_id, clean_number(affaire.get("commande_ht"))))
    affaire["pointage_vs_facturation_gap"] = round(clean_number(affaire.get("facturation_totale")) - clean_number(affaire.get("pointage_progress_amount")), 2)
    affaire["pointage_vs_financial_gap_pct"] = round((clean_number(affaire.get("taux_avancement_financier")) - clean_number(affaire.get("pointage_progress_ratio"))) * 100.0, 2)
    affaire["insights"] = FinanceService.compute_insights(affaire)
    return {"ok": True, "affaire": affaire}


@app.get("/api/project-management/board", response_class=JSONResponse)
def api_project_management_board(affaire_id: str = Query(default=""), affaire_name: str = Query(default=""), start_date: str = Query(default=""), end_date: str = Query(default="")):
    name = clean_text(affaire_name)
    if affaire_id and not name:
        cache = service.get_finance_cache()
        item = cache.get("items", {}).get(affaire_id)
        if not item:
            raise HTTPException(status_code=404, detail=f"Affaire introuvable : {affaire_id}")
        candidates = [clean_text(item.get("affaire")), clean_text(item.get("display_name"))]
        candidates = [c for i, c in enumerate(candidates) if c and c not in candidates[:i]]
        if not candidates:
            raise HTTPException(status_code=400, detail="Aucun nom projet exploitable pour cette affaire")
        boards = [metronome_service.build_project_board(c, start_date=start_date, end_date=end_date) for c in candidates]
        boards.sort(key=lambda b: (clean_number((b.get("match_debug") or {}).get("match_score", 0)), len(b.get("rows") or [])), reverse=True)
        best = boards[0]
        if isinstance(best.get("match_debug"), dict):
            best["match_debug"]["searched_variants"] = candidates
        return best
    if not name:
        raise HTTPException(status_code=400, detail="affaire_id ou affaire_name requis")
    return metronome_service.build_project_board(name, start_date=start_date, end_date=end_date)


@app.get("/api/project-management/pointage", response_class=JSONResponse)
def api_project_management_pointage(affaire_id: str = Query(default=""), affaire_name: str = Query(default="")):
    key = PointageService._project_key(affaire_id, affaire_name)
    if key.endswith("name::"):
        raise HTTPException(status_code=400, detail="affaire_id ou affaire_name requis")
    return pointage_service.get_project_data(key)


@app.post("/api/project-management/pointage/planning/import", response_class=JSONResponse)
async def api_project_management_pointage_import_planning(file: UploadFile = File(...), affaire_id: str = Query(default=""), affaire_name: str = Query(default="")):
    key = PointageService._project_key(affaire_id, affaire_name)
    if key.endswith("name::"):
        raise HTTPException(status_code=400, detail="affaire_id ou affaire_name requis")
    raw = await file.read()
    return pointage_service.import_planning(key, raw)


@app.post("/api/project-management/pointage/save", response_class=JSONResponse)
def api_project_management_pointage_save(
    payload: Dict[str, Any] = Body(default={}),
    affaire_id: str = Query(default=""),
    affaire_name: str = Query(default=""),
):
    key = PointageService._project_key(affaire_id, affaire_name)
    if key.endswith("name::"):
        raise HTTPException(status_code=400, detail="affaire_id ou affaire_name requis")
    pointage_patch = payload.get("pointage_patch", {}) if isinstance(payload, dict) else {}
    work_state = payload.get("work_state", {}) if isinstance(payload, dict) else {}
    if not isinstance(pointage_patch, dict):
        raise HTTPException(status_code=400, detail="pointage_patch invalide")
    return pointage_service.save_pointage(key, pointage_patch, work_state)


@app.post("/api/project-management/pointage/import-suivi", response_class=JSONResponse)
async def api_project_management_pointage_import_suivi(file: UploadFile = File(...), affaire_id: str = Query(default=""), affaire_name: str = Query(default="")):
    key = PointageService._project_key(affaire_id, affaire_name)
    if key.endswith("name::"):
        raise HTTPException(status_code=400, detail="affaire_id ou affaire_name requis")
    raw = await file.read()
    return pointage_service.import_suivi(key, raw)


@app.get("/api/project-management/pointage/export", response_class=JSONResponse)
def api_project_management_pointage_export(affaire_id: str = Query(default=""), affaire_name: str = Query(default="")):
    key = PointageService._project_key(affaire_id, affaire_name)
    if key.endswith("name::"):
        raise HTTPException(status_code=400, detail="affaire_id ou affaire_name requis")
    return pointage_service.export_suivi(key)


@app.get("/api/finance/affaire/{affaire_id}/export-csv")
def api_affaire_export_csv(affaire_id: str):
    cache = service.get_finance_cache()
    item = cache.get("items", {}).get(affaire_id)
    if not item:
        raise HTTPException(status_code=404, detail=f"Affaire introuvable : {affaire_id}")

    out = io.StringIO()
    writer = csv.writer(out, delimiter=";")
    writer.writerow(["Affaire", item.get("display_name", "")])
    writer.writerow(["Client", item.get("client", "")])
    writer.writerow(["Mission", item.get("affaire", "")])
    writer.writerow([])
    writer.writerow(["Mois", "Prévisionnel", "Facturé", "Écart"])
    for month in MONTHS:
        row = item.get("mensuel", {}).get(month, {})
        writer.writerow([MONTH_LABELS[month], row.get("previsionnel", 0), row.get("facture", 0), row.get("ecart", 0)])
    writer.writerow([])
    writer.writerow(["Tag", "Mission", "Numero", "💰 Commandes achetées", "🧾 Facturation totale", "⚠ Reste à facturer", "Total prévisionnel", "Total facturé"])
    for mission in item.get("missions", []):
        writer.writerow([
            mission.get("tag", ""),
            mission.get("label", ""),
            mission.get("numero", ""),
            mission.get("commande_ht", 0),
            mission.get("facturation_totale", (clean_number(mission.get("anteriorite")) + clean_number(mission.get("facture_2026", mission.get("facturation_cumulee_2026", 0))))),
            mission.get("reste_a_facturer", 0),
            mission.get("total_previsionnel", 0),
            mission.get("total_facture", 0),
        ])
    out.seek(0)
    return StreamingResponse(
        iter([out.getvalue().encode("utf-8-sig")]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=finance-{affaire_id}.csv"},
    )
